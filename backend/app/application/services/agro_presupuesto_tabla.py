"""Lector del libro de presupuesto que arma el negocio, tal como lo arma.

El archivo `AGROPECUARIA.xlsx` no tiene la forma que pide la carga masiva
genérica —una fila por meta, con su columna `dimension`—. Tiene la forma de una
tabla dinámica de Excel: los vendedores agrupados por canal, con las filas de
subtotal intercaladas, y el presupuesto partido en dos bloques de columnas
—kilos en uno, pesos en otro— que comparten la misma lista de filas.

Pedirle al negocio que reformatee su libro para poder cargarlo es la clase de
petición que acaba en «entonces lo sigo llevando en Excel». Este módulo lee el
libro que ya existe.

── Las tres trampas de este formato ─────────────────────────────────────────

**1. Las filas de subtotal.** `CALL CENTER`, `MAYORISTA`, `TAT`,
`SUPERMAYORISTA` y `Total general` son sumas de las filas que las rodean, no
miembros. Cargarlas metería el presupuesto **dos veces**: una por vendedor y
otra por su canal.

**2. El nombre no es la clave.** El libro identifica al vendedor por su nombre y
la venta llega identificada por su cédula. Guardar la meta bajo el nombre la
dejaría colgada de una clave que ninguna venta usa, y el cumplimiento de esa
persona saldría **cero para siempre sin que nada fallara**. Por eso cada nombre
se resuelve contra el catálogo que dejó la ingesta —que trae las cédulas del
origen— y lo que no resuelve se rechaza con su motivo, nunca se inventa.

**3. El mismo nombre, escrito de varias formas.** En el mismo libro conviven
`CABARCA LACHE KAREN DANIEL`, `CABRERA LACHE KAREN DANIELA` y, en la API,
`CABARCA LACHE KAREN DANIELA`. La comparación normaliza mayúsculas, tildes y
espacios repetidos, que resuelve la mayoría; lo que sigue sin casar se reporta
para que alguien lo mire, porque adivinar a quién se parece más un nombre es
justo lo que no debe hacer un sistema que reparte metas.
"""

from __future__ import annotations

import io
import unicodedata
from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from typing import Any

#: Encabezados que marcan los dos bloques dentro de la hoja. Se buscan por texto
#: y no por letra de columna: la hoja es una tabla dinámica y sus columnas se
#: mueven en cuanto alguien añade un campo.
ENCABEZADO_KILOS = "PPT EN KILO"
ENCABEZADO_PESOS = "PPT EN $"
ENCABEZADO_VENDEDOR = "VENDEDORES"

#: La precisión con la que el modelo guarda cada magnitud. Se redondea **al
#: leer** y no al comparar: si el libro trae 6.315.016.727,261173 y el sistema
#: guarda dos decimales, la cifra almacenada es 6.315.016.727,26 y es esa la que
#: hay que contrastar contra el total. Comparar la del libro sin redondear
#: dejaría un descuadre de una milésima de peso en cada carga, que no es un
#: descuadre: es una diferencia de precisión disfrazada de error.
DECIMALES_MONTO = Decimal("0.01")
DECIMALES_KILOS = Decimal("0.001")

#: Hasta qué fila se buscan los encabezados. La tabla dinámica los pone en las
#: primeras filas; recorrer la hoja entera solo daría falsos positivos.
FILAS_ENCABEZADO = 8


@dataclass(frozen=True, slots=True)
class MetaLeida:
    """Una meta leída del libro, ya resuelta contra el catálogo."""

    fila: int
    nombre: str
    clave: str
    monto: Decimal
    kilos: Decimal


@dataclass(frozen=True, slots=True)
class LecturaTabla:
    """Lo leído, lo omitido y lo que no se pudo resolver.

    `omitidas` y `sin_resolver` se devuelven por separado a propósito. Las
    primeras son las filas de subtotal, que **deben** quedar fuera y no son un
    problema; las segundas son nombres que no casaron con nadie del catálogo, y
    esas sí hay que mirarlas: puede ser un vendedor nuevo, o puede ser un
    presupuesto que se está quedando sin asignar.
    """

    metas: list[MetaLeida]
    omitidas: list[str]
    sin_resolver: list[str]
    #: El «Total general» que trae el propio libro, si lo trae. Sirve para
    #: contrastar contra la suma de lo cargado: si no coinciden, algo se quedó
    #: fuera y el aviso vale más que el dato.
    total_libro_monto: Decimal | None
    total_libro_kilos: Decimal | None


def normalizar_nombre(valor: object) -> str:
    """Mayúsculas, sin tildes y con los espacios colapsados.

    `MORA  KATIA MARGARITA` —con dos espacios, tal como está en el libro— y
    `Mora Katia Margarita` tienen que ser la misma persona.
    """
    texto = "" if valor is None else str(valor)
    sin_tildes = "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    )
    return " ".join(sin_tildes.upper().split())


def _a_decimal(valor: object) -> Decimal | None:
    """Un número de celda a `Decimal`, **sin pasar por `float`**.

    Las celdas de Excel llegan ya como `float` desde openpyxl, así que la
    conversión se hace por su representación decimal y no por el binario: es la
    diferencia entre 901.076.239,05 y 901.076.239,0499999.
    """
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        return None
    try:
        return Decimal(repr(valor) if isinstance(valor, float) else str(valor))
    except (InvalidOperation, ValueError):
        return None


def _redondear(valor: Decimal | None, paso: Decimal) -> Decimal | None:
    return None if valor is None else valor.quantize(paso, rounding=ROUND_HALF_UP)


def _localizar(hoja: Any) -> tuple[int, dict[str, int], int] | None:
    """Dónde están los encabezados: la fila, las columnas y dónde empiezan los datos.

    Devuelve `None` si la hoja no tiene esta forma, que es como el llamador
    distingue este libro de una carga masiva genérica.
    """
    for numero in range(1, FILAS_ENCABEZADO + 1):
        columnas: dict[str, int] = {}
        for celda in hoja[numero]:
            texto = normalizar_nombre(celda.value)
            if not texto:
                continue
            if texto == normalizar_nombre(ENCABEZADO_KILOS) and "kilos" not in columnas:
                columnas["kilos"] = celda.column
            elif texto == normalizar_nombre(ENCABEZADO_PESOS) and "monto" not in columnas:
                columnas["monto"] = celda.column
            elif texto == normalizar_nombre(ENCABEZADO_VENDEDOR):
                # Hay varios bloques y cada uno repite «VENDEDORES». Se guarda el
                # primero de cada bloque: el que queda justo a la izquierda de su
                # columna de cifras.
                columnas.setdefault(f"nombre_{celda.column}", celda.column)

        # Las dos columnas de cifras **y** una de nombres a la izquierda de cada
        # una. Ese tercer requisito es el que distingue esta tabla de la hoja de
        # datos crudos, que tiene los mismos dos encabezados —el presupuesto
        # repetido en cada linea de venta— y ninguna columna `VENDEDORES`.
        # Leerla a ella sumaria el presupuesto una vez por cada factura.
        if "kilos" in columnas and "monto" in columnas:
            nombres = [c for k, c in columnas.items() if k.startswith("nombre_")]
            if any(c < columnas["kilos"] for c in nombres) and any(
                c < columnas["monto"] for c in nombres
            ):
                return numero, columnas, numero + 1
    return None


def _columna_nombre(columnas: dict[str, int], columna_cifra: int) -> int:
    """La columna de nombres que le corresponde a un bloque de cifras.

    Es la de `VENDEDORES` inmediatamente a su izquierda. Buscarla por posición y
    no por un desplazamiento fijo es lo que aguanta que alguien inserte una
    columna en medio de la hoja.
    """
    candidatas = [c for k, c in columnas.items() if k.startswith("nombre_") and c < columna_cifra]
    return max(candidatas)


def leer(contenido: bytes, claves_por_nombre: dict[str, str]) -> LecturaTabla | None:
    """Lee el libro del negocio. `None` si no tiene esta forma.

    `claves_por_nombre` traduce nombre normalizado a la clave del origen —la
    cédula del vendedor—, y sale del catálogo que dejó la ingesta. Un nombre que
    no esté ahí **no se carga**: se reporta.
    """
    from openpyxl import load_workbook

    libro = load_workbook(io.BytesIO(contenido), data_only=True)
    try:
        for nombre_hoja in libro.sheetnames:
            hoja = libro[nombre_hoja]
            ubicacion = _localizar(hoja)
            if ubicacion is not None:
                return _leer_hoja(hoja, ubicacion, claves_por_nombre)
        return None
    finally:
        libro.close()


def _leer_hoja(
    hoja: Any,
    ubicacion: tuple[int, dict[str, int], int],
    claves_por_nombre: dict[str, str],
) -> LecturaTabla:
    _, columnas, primera = ubicacion
    col_kilos, col_monto = columnas["kilos"], columnas["monto"]
    col_nombre_k = _columna_nombre(columnas, col_kilos)
    col_nombre_m = _columna_nombre(columnas, col_monto)

    metas: list[MetaLeida] = []
    omitidas: list[str] = []
    sin_resolver: list[str] = []
    #: Filas sin clave, con la posición que ocupaban entre las metas resueltas.
    pendientes: list[tuple[int, str, Decimal | None, Decimal | None]] = []
    total_monto: Decimal | None = None
    total_kilos: Decimal | None = None

    for numero in range(primera, hoja.max_row + 1):
        nombre_k = normalizar_nombre(hoja.cell(numero, col_nombre_k).value)
        nombre_m = normalizar_nombre(hoja.cell(numero, col_nombre_m).value)
        if not nombre_k and not nombre_m:
            continue

        # Los dos bloques comparten la lista de filas. Si dejaran de compartirla
        # —alguien ordena uno de los dos— la meta en pesos de una persona se
        # guardaría con los kilos de otra, y nada fallaría. Se comprueba.
        if nombre_k and nombre_m and nombre_k != nombre_m:
            sin_resolver.append(
                f"fila {numero}: los bloques no coinciden ({nombre_k} / {nombre_m})"
            )
            continue

        nombre = nombre_k or nombre_m
        kilos = _redondear(_a_decimal(hoja.cell(numero, col_kilos).value), DECIMALES_KILOS)
        monto = _redondear(_a_decimal(hoja.cell(numero, col_monto).value), DECIMALES_MONTO)

        if nombre.startswith("TOTAL GENERAL"):
            total_kilos = kilos if total_kilos is None else total_kilos
            total_monto = monto if total_monto is None else total_monto
            omitidas.append(nombre)
            continue

        clave = claves_por_nombre.get(nombre)
        if clave is None:
            # Todavía no se sabe si es un subtotal de canal o un nombre que no
            # casó. Se anota y se decide al final, cuando ya se puede comprobar
            # si sus cifras son la suma de las filas que vienen debajo.
            pendientes.append((len(metas), nombre, monto, kilos))
            continue

        metas.append(
            MetaLeida(
                fila=numero,
                nombre=nombre,
                clave=clave,
                monto=monto or Decimal(0),
                kilos=kilos or Decimal(0),
            )
        )

    # Un subtotal es, literalmente, la suma de las filas de su grupo. Se
    # comprueba en vez de reconocerlo por su nombre: la lista de canales cambia
    # —hoy son cuatro, mañana cinco— y un sistema que los lleve escritos dentro
    # empezaría a cargar el nuevo canal como si fuera un vendedor, duplicando su
    # presupuesto sin que nada fallara.
    #
    # Lo que no cuadre con ninguna suma se reporta como lo que es: un nombre que
    # no se pudo cruzar, y que alguien tiene que mirar.
    for indice, (desde, nombre, monto_p, kilos_p) in enumerate(pendientes):
        hasta = pendientes[indice + 1][0] if indice + 1 < len(pendientes) else len(metas)
        grupo = metas[desde:hasta]
        if not grupo:
            sin_resolver.append(nombre)
            continue

        suma_monto = _redondear(sum((m.monto for m in grupo), Decimal(0)), DECIMALES_MONTO)
        suma_kilos = _redondear(sum((m.kilos for m in grupo), Decimal(0)), DECIMALES_KILOS)
        es_subtotal = (monto_p is None or monto_p == suma_monto) and (
            kilos_p is None or kilos_p == suma_kilos
        )
        (omitidas if es_subtotal else sin_resolver).append(nombre)

    return LecturaTabla(
        metas=metas,
        omitidas=omitidas,
        sin_resolver=sin_resolver,
        total_libro_monto=total_monto,
        total_libro_kilos=total_kilos,
    )
