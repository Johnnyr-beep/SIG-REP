"""Presupuesto: parametrización, versionado, carga masiva y cierre (§3.3, §7).

Dos reglas mandan aquí y las dos están hechas cumplir en el servicio, no en la
capa HTTP, para que sigan valiendo cuando el presupuesto se cargue desde un
script o desde un job:

1. Un período cerrado no admite cambios de presupuesto.
2. Todo cambio queda con autor, fecha y motivo.

Y una tercera, que llegó después de una auditoría: el **alcance** viaja hasta
aquí. La cifra de presupuesto de un punto de venta es información sensible
entre responsables (§8.4), así que tanto la consulta como la escritura reciben
la lista de puntos permitidos —`None` = todos— y la aplican en la consulta SQL,
no filtrando la respuesta después.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from sqlalchemy import func, select, update
from sqlalchemy.orm import Session

from app.application.services.periodos import obtener_o_crear_periodo, obtener_periodo
from app.core.errors import (
    ErrorAutorizacion,
    ErrorNoEncontrado,
    ErrorPeriodoCerrado,
    ErrorValidacion,
)
from app.core.logging import obtener_logger
from app.domain.normalizacion import (
    CATEGORIAS_RETIRADAS,
    ESCALA_DINERO,
    ESCALA_KILOS,
    clave_columna,
    destinos_de_categoria_retirada,
    normalizar_texto,
)
from app.domain.reparto import (
    DESCRIPCION_NIVEL,
    Nivel,
    elegir_pesos,
    repartir_proporcional,
)
from app.infrastructure.models.catalogo import Categoria
from app.infrastructure.models.mixins import CERO_DINERO, CERO_KILOS, ahora_utc
from app.infrastructure.models.organizacion import PuntoVenta
from app.infrastructure.models.periodo import Periodo
from app.infrastructure.models.presupuesto import Presupuesto, PresupuestoHistorial
from app.infrastructure.models.usuario import Usuario
from app.infrastructure.models.venta import VentaLinea
from app.schemas.presupuesto import (
    ErrorFila,
    HistorialSalida,
    PeriodoSalida,
    PresupuestoSalida,
    ResultadoCargaMasiva,
)

logger = obtener_logger(__name__)

#: Hoja del libro vigente de la que sale el presupuesto por PDV y categoría
#: (§3.1 y §3.3). Se compara con el nombre saneado: en este libro las hojas
#: llegan con espacios de más («CLIENTES ») más veces de las que uno esperaría.
HOJA_CUMPLIMIENTO = "CUMPLIMIENTO PPTO"

#: Encabezados admitidos en la carga masiva. Se aceptan varias grafías porque
#: el archivo lo arma el negocio a mano y exigirle una plantilla exacta es
#: garantizar que la primera carga falle.
_ALIAS_COLUMNAS: dict[str, str] = {
    "punto_venta": "punto_venta",
    "punto de venta": "punto_venta",
    "pdv": "punto_venta",
    "c.o.": "punto_venta",
    "co": "punto_venta",
    "codigo_co": "punto_venta",
    "categoria": "categoria",
    "categoría": "categoria",
    "nueva categoria": "categoria",
    "monto": "monto",
    "ppto": "monto",
    "presupuesto": "monto",
    "kilos": "kilos",
    "ppto en kilo": "kilos",
    "ppto kilos": "kilos",
}


def motivo_categoria_retirada(nombre: str, destinos: Sequence[str]) -> str:
    """El motivo con el que se rechaza una fila de una categoría que ya no existe.

    Tiene que ser **accionable**, no solo cierto. `OTROS` desapareció y el libro
    que el negocio usa hoy todavía trae ese renglón con 616 000 000 de los
    20 000 000 000 del presupuesto de la compañía. Esas filas no se pueden
    cargar, y hay tres salidas posibles de las cuales dos son inaceptables:
    descartarlas en silencio —el consolidado quedaría 616 millones corto y el
    cumplimiento de la compañía saldría inflado sin que nadie sepa por qué— o
    repartirlas por cuenta del sistema con un criterio inventado. Queda la
    tercera: rechazarlas diciendo exactamente qué hay que hacer.
    """
    reparto = ", ".join(destinos[:-1]) + " y " + destinos[-1] if len(destinos) > 1 else destinos[0]
    return (
        f"La categoría {nombre} ya no existe: SIGREP usa las categorías reales de SIESA. "
        f"Reparta ese presupuesto entre {reparto} en la fila que corresponda del archivo, "
        "o captúrelo por pantalla en Presupuesto. El sistema no lo reparte solo porque el "
        "criterio del reparto lo decide el negocio."
    )


def motivo_reparto(
    codigo_periodo: str,
    nombre_categoria: str,
    nivel_monto: Nivel,
    nivel_kilos: Nivel,
) -> str:
    """El motivo con el que queda historiado cada renglón del reparto (§3.3).

    Tiene que decir tres cosas y las dice: **de dónde viene el número** (reparto
    proporcional a la venta del período, y con qué base se hizo cada magnitud),
    **de dónde salía** (el presupuesto de la categoría retirada) y **qué vale**
    (una cifra de partida que la gerencia tiene que confirmar). Un historial que
    solo dijera «reparto automático» obligaría a reconstruir a mano, dentro de
    seis meses, si esos 4 300 000 de HUEVOS los decidió alguien o los calculó una
    división.

    Cabe en `presupuesto_historial.motivo`, que es `String(400)`; la prueba
    `test_el_motivo_cabe_en_la_columna` lo fija para que nadie lo alargue sin
    darse cuenta.
    """
    if nivel_monto == nivel_kilos:
        base = f"Base del reparto: {DESCRIPCION_NIVEL[nivel_monto]}."
    else:
        base = (
            f"Base del reparto: monto, {DESCRIPCION_NIVEL[nivel_monto]}; "
            f"kilos, {DESCRIPCION_NIVEL[nivel_kilos]}."
        )
    return (
        f"Reparto proporcional a la venta del período {codigo_periodo} del presupuesto de "
        f"«{nombre_categoria}», retirada. {base} Cifra de partida: la gerencia debe confirmarla."
    )


@dataclass(frozen=True, slots=True)
class FilaCarga:
    """Una fila ya normalizada del archivo de carga masiva."""

    numero: int
    codigo_punto_venta: str
    nombre_categoria: str
    monto: Decimal
    kilos: Decimal


@dataclass(frozen=True, slots=True)
class ParteReparto:
    """Lo que una categoría destino recibe en un punto de venta."""

    categoria: str
    monto: Decimal
    kilos: Decimal


@dataclass(frozen=True, slots=True)
class RepartoPunto:
    """El reparto completo del presupuesto de un punto de venta."""

    punto_venta: str
    monto_origen: Decimal
    kilos_origen: Decimal
    nivel_monto: Nivel
    nivel_kilos: Nivel
    partes: tuple[ParteReparto, ...]

    @property
    def monto_repartido(self) -> Decimal:
        return sum((p.monto for p in self.partes), CERO_DINERO)

    @property
    def kilos_repartidos(self) -> Decimal:
        return sum((p.kilos for p in self.partes), CERO_KILOS)


@dataclass(frozen=True, slots=True)
class ResultadoReparto:
    """Lo que hizo —o haría, en simulación— una pasada del reparto."""

    periodo: str
    categoria_retirada: str
    destinos: tuple[str, ...]
    simulacion: bool
    puntos: tuple[RepartoPunto, ...]

    @property
    def monto_origen(self) -> Decimal:
        return sum((p.monto_origen for p in self.puntos), CERO_DINERO)

    @property
    def kilos_origen(self) -> Decimal:
        return sum((p.kilos_origen for p in self.puntos), CERO_KILOS)

    @property
    def monto_repartido(self) -> Decimal:
        return sum((p.monto_repartido for p in self.puntos), CERO_DINERO)

    @property
    def kilos_repartidos(self) -> Decimal:
        return sum((p.kilos_repartidos for p in self.puntos), CERO_KILOS)

    @property
    def cuadra(self) -> bool:
        """Lo único que no se negocia: no se pierde ni se inventa un peso."""
        return (
            self.monto_repartido == self.monto_origen and self.kilos_repartidos == self.kilos_origen
        )


class PresupuestoService:
    """Casos de uso del presupuesto."""

    def __init__(self, sesion: Session) -> None:
        self._sesion = sesion

    # ── Consulta ──────────────────────────────────────────────────────────────

    def listar(
        self,
        codigo_periodo: str,
        codigo_punto_venta: str | None = None,
        alcance: list[int] | None = None,
    ) -> list[PresupuestoSalida]:
        """Presupuesto del período dentro del alcance del usuario.

        `alcance=None` significa «todos los puntos». Una lista vacía significa
        «ninguno» y devuelve vacío: un JEFE_PDV al que nadie le configuró sus
        puntos no ve el presupuesto de la compañía por omisión.
        """
        periodo = obtener_periodo(self._sesion, codigo_periodo)
        consulta = (
            select(Presupuesto)
            .join(PuntoVenta, Presupuesto.punto_venta_id == PuntoVenta.id)
            .join(Categoria, Presupuesto.categoria_id == Categoria.id)
            .where(Presupuesto.periodo_id == periodo.id)
            .order_by(PuntoVenta.codigo_co, Categoria.orden)
        )
        if codigo_punto_venta:
            consulta = consulta.where(PuntoVenta.codigo_co == codigo_punto_venta.strip())
        if alcance is not None:
            # `or [-1]` mantiene la consulta válida con alcance vacío y devuelve
            # cero filas, igual que en `reportes_service`.
            consulta = consulta.where(PuntoVenta.id.in_(alcance or [-1]))

        autores = self._autores()
        return [
            PresupuestoSalida(
                punto_venta=p.punto_venta.codigo_co,
                categoria=p.categoria.nombre,
                monto=p.monto,
                kilos=p.kilos,
                actualizado_en=p.actualizado_en,
                actualizado_por=autores.get(p.actualizado_por_id),
            )
            for p in self._sesion.execute(consulta).scalars()
        ]

    def historial(
        self,
        codigo_periodo: str | None = None,
        codigo_punto_venta: str | None = None,
        alcance: list[int] | None = None,
    ) -> list[HistorialSalida]:
        """Rastro de cambios dentro del alcance del usuario.

        El historial lleva los mismos importes que el presupuesto —de hecho
        lleva el anterior *y* el nuevo—, así que se filtra exactamente igual.
        """
        consulta = select(PresupuestoHistorial).order_by(PresupuestoHistorial.cuando.desc())
        if codigo_periodo:
            periodo = obtener_periodo(self._sesion, codigo_periodo)
            consulta = consulta.where(PresupuestoHistorial.periodo_id == periodo.id)
        if codigo_punto_venta:
            punto = self._punto_por_codigo(codigo_punto_venta)
            consulta = consulta.where(PresupuestoHistorial.punto_venta_id == punto.id)
        if alcance is not None:
            consulta = consulta.where(PresupuestoHistorial.punto_venta_id.in_(alcance or [-1]))

        autores = self._autores()
        return [
            HistorialSalida(
                cuando=h.cuando,
                quien=autores.get(h.usuario_id),
                campo=h.campo,
                valor_anterior=h.valor_anterior,
                valor_nuevo=h.valor_nuevo,
                motivo=h.motivo,
            )
            for h in self._sesion.execute(consulta).scalars()
        ]

    # ── Parametrización ───────────────────────────────────────────────────────

    def guardar(
        self,
        *,
        codigo_periodo: str,
        punto_venta_id: int,
        categoria_id: int,
        monto: Decimal,
        kilos: Decimal,
        motivo: str,
        usuario: Usuario | None = None,
        alcance: list[int] | None = None,
    ) -> PresupuestoSalida:
        """Fija el presupuesto de una celda y deja el rastro del cambio.

        Solo se historian los campos que **cambiaron de valor**: guardar sin
        modificar nada no ensucia el historial, que es lo que hace que el
        historial siga siendo legible al cabo de un año.

        `alcance` es la lista de puntos que quien escribe tiene asignados
        (`None` = todos). Se comprueba aquí y no solo en el router porque la
        carga masiva entra por este mismo método fila a fila.
        """
        periodo = obtener_o_crear_periodo(self._sesion, codigo_periodo)
        self._exigir_periodo_abierto(periodo)

        punto = self._sesion.get(PuntoVenta, punto_venta_id)
        if punto is None:
            raise ErrorNoEncontrado(f"No existe el punto de venta {punto_venta_id}.")
        self._exigir_alcance(punto, alcance)
        categoria = self._sesion.get(Categoria, categoria_id)
        if categoria is None:
            raise ErrorNoEncontrado(f"No existe la categoría {categoria_id}.")

        if not punto.presupuestado:
            raise ErrorValidacion(
                f"El punto de venta {punto.codigo_co} {punto.nombre} no se presupuesta. "
                "Su venta se reporta aparte."
            )

        fila = self._sesion.execute(
            select(Presupuesto).where(
                Presupuesto.periodo_id == periodo.id,
                Presupuesto.punto_venta_id == punto_venta_id,
                Presupuesto.categoria_id == categoria_id,
            )
        ).scalar_one_or_none()

        anterior_monto = fila.monto if fila is not None else None
        anterior_kilos = fila.kilos if fila is not None else None

        if fila is None:
            fila = Presupuesto(
                periodo_id=periodo.id,
                punto_venta_id=punto_venta_id,
                categoria_id=categoria_id,
                monto=monto,
                kilos=kilos,
            )
            self._sesion.add(fila)
        else:
            fila.monto = monto
            fila.kilos = kilos

        fila.actualizado_por_id = usuario.id if usuario else None
        self._sesion.flush()

        self._historiar(fila, "monto", anterior_monto, monto, motivo, usuario)
        self._historiar(fila, "kilos", anterior_kilos, kilos, motivo, usuario)
        self._sesion.flush()

        logger.info(
            "presupuesto_guardado",
            periodo=periodo.codigo,
            punto_venta=punto.codigo_co,
            categoria=categoria.nombre,
        )
        return PresupuestoSalida(
            punto_venta=punto.codigo_co,
            categoria=categoria.nombre,
            monto=fila.monto,
            kilos=fila.kilos,
            actualizado_en=fila.actualizado_en,
            actualizado_por=usuario.usuario if usuario else None,
        )

    # ── Reparto de una categoría retirada ─────────────────────────────────────

    def repartir_categoria_retirada(
        self,
        *,
        codigo_periodo: str,
        nombre_categoria: str,
        destinos: Sequence[str] | None = None,
        usuario: Usuario | None = None,
        simulacion: bool = False,
    ) -> ResultadoReparto:
        """Reparte el presupuesto de una categoría retirada y la deja vacía.

        Es la ejecución de una decisión que ya tomó el negocio: `OTROS` se
        retira y sus 616 000 000 se reparten **a prorrata de la venta real ya
        cargada** entre las cuatro categorías que la sustituyen. El sistema no
        elige el criterio; lo aplica y deja dicho en el historial que la cifra es
        un punto de partida a confirmar por pantalla.

        Cuatro reglas mandan aquí:

        1. **El reparto es por punto de venta.** Cada uno de los 15 PDV tiene su
           propio monto en la categoría retirada y se reparte con **la venta de
           ese PDV**. Usar la proporción global le pondría a MALAMBO el perfil de
           consumo de BUCARAMANGA, que es exactamente el error que hace inútil un
           presupuesto por punto.
        2. **El monto sigue la venta en pesos; los kilos, la venta en kilos.**
           No es un adorno: DOMICILIOS pesa el 6 % de la venta en dinero de las
           cuatro categorías y el 19 % en kilos. Repartir los kilos con la
           proporción del dinero le daría a DOMICILIOS un presupuesto en kilos
           que no tiene nada que ver con lo que mueve. Cada magnitud se reparte
           con su propia magnitud, y la cascada de `elegir_pesos` se aplica a
           cada una por separado —de ahí que el motivo pueda nombrar dos bases
           distintas—.
        3. **No se pierde ni se inventa un peso.** Ver `domain/reparto.py`.
        4. **Todo queda historiado** (§3.3) y **un período cerrado se rechaza**
           (§7), las dos por `guardar`, que es por donde pasa cada renglón.

        Al terminar, las filas de presupuesto de la categoría retirada **se
        borran**: es lo que permite que la migración `0005` elimine la categoría
        sin destruir nada. Antes de borrarlas se historia su vaciado
        (`monto → 0`) y se desliga el historial que las apuntaba —
        `presupuesto_historial.presupuesto_id` es anulable justo para esto, ver
        el modelo—, de modo que el rastro sobrevive a la fila.

        Es **idempotente**: una segunda pasada no encuentra filas que repartir y
        devuelve un resultado vacío sin tocar nada.

        Con `simulacion=True` calcula y devuelve el reparto sin escribir. Es lo
        que imprime `--simular` en la línea de órdenes, y es la forma de que
        alguien mire los números **antes** de moverlos.
        """
        periodo = obtener_periodo(self._sesion, codigo_periodo)
        self._exigir_periodo_abierto(periodo)

        retirada = self._categoria_exacta(nombre_categoria)
        nombres_destino = tuple(destinos or destinos_de_categoria_retirada(retirada.nombre) or ())
        if not nombres_destino:
            raise ErrorValidacion(
                f"No se sabe entre qué categorías repartir «{retirada.nombre}». "
                "Indíquelas explícitamente o dé de alta la categoría retirada en "
                "CATEGORIAS_RETIRADAS."
            )
        categorias_destino = [self._categoria_exacta(nombre) for nombre in nombres_destino]
        if retirada.id in {categoria.id for categoria in categorias_destino}:
            raise ErrorValidacion(f"«{retirada.nombre}» no puede ser destino de su propio reparto.")

        filas = list(
            self._sesion.execute(
                select(Presupuesto)
                .join(PuntoVenta, Presupuesto.punto_venta_id == PuntoVenta.id)
                .where(
                    Presupuesto.periodo_id == periodo.id,
                    Presupuesto.categoria_id == retirada.id,
                )
                .order_by(PuntoVenta.codigo_co)
            ).scalars()
        )

        ids_destino = [categoria.id for categoria in categorias_destino]
        venta = self._venta_por_punto_y_categoria(periodo.id, ids_destino)
        globales = self._pesos_globales(venta, ids_destino)

        repartos = [
            self._repartir_una_fila(
                fila=fila,
                periodo_codigo=codigo_periodo,
                retirada=retirada,
                categorias_destino=categorias_destino,
                venta=venta,
                globales=globales,
                usuario=usuario,
                simulacion=simulacion,
            )
            for fila in filas
        ]

        resultado = ResultadoReparto(
            periodo=codigo_periodo,
            categoria_retirada=retirada.nombre,
            destinos=nombres_destino,
            simulacion=simulacion,
            puntos=tuple(repartos),
        )
        if not resultado.cuadra:  # pragma: no cover - lo impide `repartir_proporcional`
            raise ErrorValidacion(
                "El reparto no cuadra: se repartieron "
                f"{resultado.monto_repartido} de {resultado.monto_origen}. No se aplica."
            )

        logger.info(
            "reparto_categoria_retirada",
            periodo=codigo_periodo,
            categoria=retirada.nombre,
            puntos=len(repartos),
            monto=str(resultado.monto_origen),
            simulacion=simulacion,
        )
        return resultado

    @staticmethod
    def _pesos_globales(
        venta: dict[tuple[int, int], tuple[Decimal, Decimal]],
        ids_destino: Sequence[int],
    ) -> tuple[list[Decimal], list[Decimal]]:
        """Venta de todo el período por categoría destino: el segundo escalón."""
        monto = [
            sum((v[0] for (_, c), v in venta.items() if c == categoria), CERO_DINERO)
            for categoria in ids_destino
        ]
        kilos = [
            sum((v[1] for (_, c), v in venta.items() if c == categoria), CERO_KILOS)
            for categoria in ids_destino
        ]
        return monto, kilos

    def _repartir_una_fila(
        self,
        *,
        fila: Presupuesto,
        periodo_codigo: str,
        retirada: Categoria,
        categorias_destino: Sequence[Categoria],
        venta: dict[tuple[int, int], tuple[Decimal, Decimal]],
        globales: tuple[list[Decimal], list[Decimal]],
        usuario: Usuario | None,
        simulacion: bool,
    ) -> RepartoPunto:
        """El reparto del presupuesto de **un** punto de venta."""
        punto_id = fila.punto_venta_id
        del_punto_monto = [
            venta.get((punto_id, categoria.id), (CERO_DINERO, CERO_KILOS))[0]
            for categoria in categorias_destino
        ]
        del_punto_kilos = [
            venta.get((punto_id, categoria.id), (CERO_DINERO, CERO_KILOS))[1]
            for categoria in categorias_destino
        ]

        pesos_monto, nivel_monto = elegir_pesos(del_punto_monto, globales[0])
        pesos_kilos, nivel_kilos = elegir_pesos(del_punto_kilos, globales[1])

        monto_origen, kilos_origen = fila.monto, fila.kilos
        partes_monto = repartir_proporcional(monto_origen, pesos_monto, ESCALA_DINERO)
        partes_kilos = repartir_proporcional(kilos_origen, pesos_kilos, ESCALA_KILOS)

        motivo = motivo_reparto(periodo_codigo, retirada.nombre, nivel_monto, nivel_kilos)

        if not simulacion:
            for categoria, monto, kilos in zip(
                categorias_destino, partes_monto, partes_kilos, strict=True
            ):
                if monto == 0 and kilos == 0:
                    # Un destino que no vendió nada recibe cero, y cero no se
                    # escribe: crearía una fila de presupuesto vacía y dos
                    # renglones de historial que no cuentan nada. El resultado
                    # devuelto sí lleva la parte en cero, que es donde hay que
                    # poder verla.
                    continue
                actual = self._fila_presupuesto(fila.periodo_id, punto_id, categoria.id)
                self.guardar(
                    codigo_periodo=periodo_codigo,
                    punto_venta_id=punto_id,
                    categoria_id=categoria.id,
                    # **Se suma a lo que ya hubiera.** Reemplazar destruiría el
                    # presupuesto que alguien capturó antes en QUESO Y LACTEOS.
                    monto=(actual.monto if actual is not None else CERO_DINERO) + monto,
                    kilos=(actual.kilos if actual is not None else CERO_KILOS) + kilos,
                    motivo=motivo,
                    usuario=usuario,
                    alcance=None,
                )
            self._vaciar_y_borrar(fila, motivo, usuario)

        return RepartoPunto(
            punto_venta=fila.punto_venta.codigo_co,
            monto_origen=monto_origen,
            kilos_origen=kilos_origen,
            nivel_monto=nivel_monto,
            nivel_kilos=nivel_kilos,
            partes=tuple(
                ParteReparto(categoria=categoria.nombre, monto=monto, kilos=kilos)
                for categoria, monto, kilos in zip(
                    categorias_destino, partes_monto, partes_kilos, strict=True
                )
            ),
        )

    def _vaciar_y_borrar(self, fila: Presupuesto, motivo: str, usuario: Usuario | None) -> None:
        """Historia el vaciado de la fila de origen y la borra.

        El orden importa. Primero se historia (`19 551 895,23 → 0`), porque ese
        renglón es la contrapartida de los cuatro que acaban de entrar y sin él
        el historial contaría cuatro altas sin ninguna baja. Después se anula el
        `presupuesto_id` de **todo** el historial que apuntaba a esta fila —el
        recién escrito y el que ya existía de la carga inicial— para que el
        `DELETE` no choque con la clave foránea. El historial sobrevive: sus
        claves de período, punto de venta y categoría son las que lo hacen
        legible, y el modelo lo dice desde el primer día.
        """
        self._historiar(fila, "monto", fila.monto, CERO_DINERO, motivo, usuario)
        self._historiar(fila, "kilos", fila.kilos, CERO_KILOS, motivo, usuario)
        self._sesion.flush()

        self._sesion.execute(
            update(PresupuestoHistorial)
            .where(PresupuestoHistorial.presupuesto_id == fila.id)
            .values(presupuesto_id=None)
        )
        self._sesion.delete(fila)
        self._sesion.flush()

    def _venta_por_punto_y_categoria(
        self, periodo_id: int, ids_categoria: Sequence[int]
    ) -> dict[tuple[int, int], tuple[Decimal, Decimal]]:
        """Venta del período agregada por (punto de venta, categoría).

        Una sola consulta agrupada, no una por PDV: son 15 puntos × 4 categorías
        y `venta_lineas` tiene 131 819 filas para nueve días. El índice
        `ix_venta_periodo_pdv_categoria` cubre exactamente este `WHERE` y este
        `GROUP BY`.
        """
        if not ids_categoria:
            return {}
        filas = self._sesion.execute(
            select(
                VentaLinea.punto_venta_id,
                VentaLinea.categoria_id,
                func.sum(VentaLinea.valor_subtotal),
                func.sum(VentaLinea.cantidad_inv),
            )
            .where(
                VentaLinea.periodo_id == periodo_id,
                VentaLinea.categoria_id.in_(ids_categoria),
            )
            .group_by(VentaLinea.punto_venta_id, VentaLinea.categoria_id)
        ).all()
        return {
            (int(punto), int(categoria)): (
                Decimal(monto or 0),
                Decimal(kilos or 0),
            )
            for punto, categoria, monto, kilos in filas
        }

    def _fila_presupuesto(
        self, periodo_id: int, punto_venta_id: int, categoria_id: int
    ) -> Presupuesto | None:
        return self._sesion.execute(
            select(Presupuesto).where(
                Presupuesto.periodo_id == periodo_id,
                Presupuesto.punto_venta_id == punto_venta_id,
                Presupuesto.categoria_id == categoria_id,
            )
        ).scalar_one_or_none()

    def _categoria_exacta(self, nombre: str) -> Categoria:
        """Categoría por nombre exacto, **incluidas las desactivadas**.

        `_categoria_por_nombre` sirve a la carga masiva y traduce una categoría
        retirada en un error accionable. Aquí hace falta lo contrario: encontrar
        precisamente la retirada, que es la que hay que vaciar.
        """
        limpio = (normalizar_texto(nombre) or "").upper()
        categoria = self._sesion.execute(
            select(Categoria).where(Categoria.nombre == limpio)
        ).scalar_one_or_none()
        if categoria is None:
            raise ErrorNoEncontrado(f"No existe la categoría {nombre!r}.")
        return categoria

    # ── Carga masiva ──────────────────────────────────────────────────────────

    def cargar_masivo(
        self,
        contenido: bytes,
        nombre_archivo: str,
        *,
        codigo_periodo: str,
        motivo: str,
        usuario: Usuario | None = None,
        alcance: list[int] | None = None,
    ) -> ResultadoCargaMasiva:
        """Carga un `.xlsx` o `.csv` con el presupuesto del período.

        Hoy el presupuesto se arma en Excel, así que la carga masiva no es una
        comodidad: es la forma en que este dato va a entrar al sistema.

        Una fila mala **no aborta la carga**. Se acepta lo que se pueda y se
        devuelve el detalle de lo rechazado con su número de fila; abortar
        entero por un typo en la fila 90 obligaría a repetir 89 filas buenas.

        El archivo lo trae el usuario, así que sus filas no son más de fiar que
        cualquier otra entrada: una fila de un punto fuera de su alcance se
        rechaza **con su motivo**, como cualquier otro error de fila. Rechazarla
        en silencio dejaría a quien carga creyendo que ese presupuesto quedó
        puesto.

        Lo mismo vale, y con más razón, para las filas de una categoría
        **retirada**: el libro vigente trae `OTROS` con 616 000 000 de los
        20 000 000 000 del presupuesto de la compañía, y esas filas salen
        rechazadas con el reparto propuesto en el motivo. Se pierde la carga de
        esas filas, no el dato: sigue en el archivo y sale en esta respuesta con
        su número de fila.
        """
        periodo = obtener_o_crear_periodo(self._sesion, codigo_periodo)
        self._exigir_periodo_abierto(periodo)

        filas, errores = self._leer_archivo(contenido, nombre_archivo)

        aceptadas = 0
        for fila in filas:
            try:
                punto = self._punto_por_codigo(fila.codigo_punto_venta)
                categoria = self._categoria_por_nombre(fila.nombre_categoria)
                self.guardar(
                    codigo_periodo=codigo_periodo,
                    punto_venta_id=punto.id,
                    categoria_id=categoria.id,
                    monto=fila.monto,
                    kilos=fila.kilos,
                    motivo=motivo,
                    usuario=usuario,
                    alcance=alcance,
                )
            except (ErrorAutorizacion, ErrorNoEncontrado, ErrorValidacion) as exc:
                errores.append(ErrorFila(fila=fila.numero, motivo=exc.mensaje))
            else:
                aceptadas += 1

        logger.info(
            "carga_masiva_presupuesto",
            periodo=codigo_periodo,
            aceptadas=aceptadas,
            rechazadas=len(errores),
        )
        return ResultadoCargaMasiva(aceptadas=aceptadas, rechazadas=len(errores), errores=errores)

    def _leer_archivo(
        self, contenido: bytes, nombre_archivo: str
    ) -> tuple[list[FilaCarga], list[ErrorFila]]:
        nombre = nombre_archivo.lower()
        if nombre.endswith(".csv"):
            crudas = self._filas_csv(contenido)
        elif nombre.endswith((".xlsx", ".xlsm")):
            crudas = self._filas_excel(contenido)
        else:
            raise ErrorValidacion("Formato no admitido. Use .xlsx o .csv.")

        filas: list[FilaCarga] = []
        errores: list[ErrorFila] = []
        for numero, cruda in crudas:
            try:
                filas.append(self._normalizar_fila(numero, cruda))
            except ErrorValidacion as exc:
                errores.append(ErrorFila(fila=numero, motivo=exc.mensaje))
        return filas, errores

    @staticmethod
    def _filas_csv(contenido: bytes) -> list[tuple[int, dict[str, object]]]:
        texto = contenido.decode("utf-8-sig", errors="replace")
        lector = csv.DictReader(io.StringIO(texto))
        # `numero` cuenta desde 2 porque la 1 es el encabezado: así el número
        # que se le devuelve al usuario coincide con el que ve en su editor.
        return [
            (numero, {k: v for k, v in fila.items() if k is not None})
            for numero, fila in enumerate(lector, start=2)
        ]

    def _filas_excel(self, contenido: bytes) -> list[tuple[int, dict[str, object]]]:
        """Filas crudas de un `.xlsx`, por el camino que corresponda.

        Hay dos formatos y los dos son legítimos: la plantilla plana —una fila
        por celda de presupuesto, con encabezado— y **la hoja
        `CUMPLIMIENTO PPTO` del libro que el negocio ya tiene**, que es de donde
        va a salir la primera carga real. Se detecta la segunda y se lee con su
        propio parser; pedirle al usuario que reordene su Excel a mano antes de
        cargarlo es la forma más segura de que la carga masiva no se use nunca.
        """
        from openpyxl import load_workbook

        libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        try:
            hoja_presupuesto = next(
                (
                    libro[nombre]
                    for nombre in libro.sheetnames
                    if str(nombre).strip().upper() == HOJA_CUMPLIMIENTO
                ),
                None,
            )
            if hoja_presupuesto is not None:
                return self._filas_hoja_cumplimiento(hoja_presupuesto)

            hoja = libro.active
            if hoja is None:  # pragma: no cover - un libro siempre trae una hoja
                raise ErrorValidacion("El archivo no tiene ninguna hoja legible.")

            iterador = hoja.iter_rows(values_only=True)
            try:
                encabezados = [str(c).strip() if c is not None else "" for c in next(iterador)]
            except StopIteration as exc:
                raise ErrorValidacion("El archivo está vacío.") from exc

            filas: list[tuple[int, dict[str, object]]] = []
            for numero, valores in enumerate(iterador, start=2):
                if all(v is None for v in valores):
                    continue
                filas.append((numero, dict(zip(encabezados, valores, strict=False))))
            return filas
        finally:
            libro.close()

    def _filas_hoja_cumplimiento(self, hoja: Any) -> list[tuple[int, dict[str, object]]]:
        """Lee la hoja `CUMPLIMIENTO PPTO` del libro vigente (§3.1, §3.3).

        La hoja no empieza por el encabezado: arriba lleva el bloque de días
        hábiles por zona, y la tabla arranca en la fila cuyo primer título es
        `CODIGO`. De ahí para abajo vienen tres clases de fila mezcladas:

            001  GRUPO 1      ← total del grupo
            402  MALAMBO      ← total del punto de venta
            402  RES          ← presupuesto de (punto de venta, categoría)  ✔
            402  CERDO        ← ídem                                        ✔

        **Solo la tercera clase se carga.** Las otras dos son sumas y el
        presupuesto del PDV y del grupo se calculan, no se capturan por
        duplicado (§3.3): cargarlas multiplicaría por dos el presupuesto de la
        compañía. La regla para distinguirlas es doble —la etiqueta es una
        categoría conocida, o al menos no es la primera fila de su bloque de
        código—, y así una categoría nueva que el negocio añada mañana no se
        pierde en silencio: llega hasta el validador y sale como error de fila
        con su motivo.

        «Categoría conocida» incluye a propósito las **retiradas**. El libro
        vigente trae el renglón `OTROS` y esa fila tiene que llegar entera al
        validador para salir con el motivo que explica el reparto; si además
        fuera la primera de su bloque —un PDV cuyo presupuesto solo tuviera esa
        categoría—, sin esta salvedad se confundiría con un total y se
        descartaría en silencio, que es justo lo que no puede pasar con 616
        millones de presupuesto.
        """
        categorias = {
            nombre.upper() for nombre in self._sesion.execute(select(Categoria.nombre)).scalars()
        }
        categorias |= set(CATEGORIAS_RETIRADAS)

        columnas: dict[str, int] = {}
        filas: list[tuple[int, dict[str, object]]] = []
        codigo_bloque: str | None = None

        for numero, valores in enumerate(hoja.iter_rows(values_only=True), start=1):
            if not valores:
                continue
            if not columnas:
                if _clave(valores[0]) == "codigo":
                    columnas = {}
                    for indice, titulo in enumerate(valores):
                        clave = _clave(titulo)
                        # Los encabezados `%` se repiten cuatro veces en la hoja;
                        # manda la primera aparición de cada nombre.
                        if clave and clave not in columnas:
                            columnas[clave] = indice
                    for obligatoria in ("codigo", "pdv", "ppto"):
                        if obligatoria not in columnas:
                            raise ErrorValidacion(
                                f"La hoja {HOJA_CUMPLIMIENTO} no trae la columna "
                                f"«{obligatoria.upper()}»."
                            )
                continue

            codigo = _celda(valores, columnas, "codigo")
            etiqueta = _texto_celda(_celda(valores, columnas, "pdv"))
            if codigo is None or etiqueta is None:
                continue

            codigo_normalizado = _texto_celda(codigo) or ""
            es_primera_del_bloque = codigo_normalizado != codigo_bloque
            codigo_bloque = codigo_normalizado
            if es_primera_del_bloque and etiqueta.upper() not in categorias:
                continue  # total del grupo o del punto de venta

            filas.append(
                (
                    numero,
                    {
                        "punto_venta": codigo,
                        "categoria": etiqueta,
                        "monto": _celda(valores, columnas, "ppto"),
                        "kilos": _celda(valores, columnas, "ppto en kilo"),
                    },
                )
            )
        if not columnas:
            raise ErrorValidacion(
                f"La hoja {HOJA_CUMPLIMIENTO} no tiene la fila de encabezado que empieza "
                "por «CODIGO»."
            )
        return filas

    @staticmethod
    def _normalizar_fila(numero: int, cruda: dict[str, object]) -> FilaCarga:
        normalizada: dict[str, object] = {}
        for clave, valor in cruda.items():
            destino = _ALIAS_COLUMNAS.get(str(clave).strip().lower())
            if destino is not None:
                normalizada[destino] = valor

        faltantes = {"punto_venta", "categoria"} - normalizada.keys()
        if faltantes:
            raise ErrorValidacion(
                "Faltan columnas obligatorias: " + ", ".join(sorted(faltantes)) + "."
            )

        codigo = str(normalizada["punto_venta"] or "").strip()
        # El C.O. llega tanto como '606' como 606; se normaliza a tres
        # posiciones con ceros a la izquierda, igual que en la ingesta (§3.4).
        if codigo.endswith(".0"):
            codigo = codigo[:-2]
        codigo = codigo.zfill(3)
        if not codigo.strip("0"):
            raise ErrorValidacion("La fila no indica punto de venta.")

        categoria = str(normalizada["categoria"] or "").strip()
        if not categoria:
            raise ErrorValidacion("La fila no indica categoría.")

        return FilaCarga(
            numero=numero,
            codigo_punto_venta=codigo,
            nombre_categoria=categoria,
            monto=_a_decimal(normalizada.get("monto"), "monto"),
            kilos=_a_decimal(normalizada.get("kilos"), "kilos"),
        )

    # ── Períodos ──────────────────────────────────────────────────────────────

    def listar_periodos(self) -> list[PeriodoSalida]:
        periodos = self._sesion.execute(
            select(Periodo).order_by(Periodo.anio.desc(), Periodo.mes.desc())
        ).scalars()
        autores = self._autores()
        return [
            PeriodoSalida(
                periodo=p.codigo,
                cerrado=p.cerrado,
                cerrado_por=autores.get(p.cerrado_por_id),
                cerrado_en=p.cerrado_en,
            )
            for p in periodos
        ]

    def cerrar_periodo(self, codigo_periodo: str, usuario: Usuario) -> PeriodoSalida:
        """Cierra el período. A partir de aquí el presupuesto es inmutable."""
        periodo = obtener_periodo(self._sesion, codigo_periodo)
        if periodo.cerrado:
            raise ErrorPeriodoCerrado(f"El período {periodo.codigo} ya estaba cerrado.")

        periodo.cerrado = True
        periodo.cerrado_por_id = usuario.id
        periodo.cerrado_en = ahora_utc()
        self._sesion.flush()

        logger.info("periodo_cerrado", periodo=periodo.codigo, usuario=usuario.usuario)
        return PeriodoSalida(
            periodo=periodo.codigo,
            cerrado=True,
            cerrado_por=usuario.usuario,
            cerrado_en=periodo.cerrado_en,
        )

    # ── Interno ───────────────────────────────────────────────────────────────

    @staticmethod
    def _exigir_alcance(punto: PuntoVenta, alcance: list[int] | None) -> None:
        """Corta la escritura sobre un punto que no es del usuario."""
        if alcance is not None and punto.id not in alcance:
            raise ErrorAutorizacion(
                f"No tiene alcance sobre el punto de venta {punto.codigo_co} {punto.nombre}."
            )

    @staticmethod
    def _exigir_periodo_abierto(periodo: Periodo) -> None:
        if periodo.cerrado:
            raise ErrorPeriodoCerrado(
                f"El período {periodo.codigo} está cerrado y no admite cambios de presupuesto."
            )

    def _historiar(
        self,
        fila: Presupuesto,
        campo: str,
        anterior: Decimal | None,
        nuevo: Decimal,
        motivo: str,
        usuario: Usuario | None,
    ) -> None:
        if anterior is not None and anterior == nuevo:
            return
        self._sesion.add(
            PresupuestoHistorial(
                presupuesto_id=fila.id,
                periodo_id=fila.periodo_id,
                punto_venta_id=fila.punto_venta_id,
                categoria_id=fila.categoria_id,
                campo=campo,
                valor_anterior=anterior,
                valor_nuevo=nuevo,
                motivo=motivo,
                usuario_id=usuario.id if usuario else None,
                cuando=self._instante(),
            )
        )

    @staticmethod
    def _instante() -> datetime:
        return ahora_utc()

    def _autores(self) -> dict[int | None, str]:
        """Mapa `id → usuario` para resolver los nombres de una sola consulta."""
        filas = self._sesion.execute(select(Usuario.id, Usuario.usuario)).all()
        return {fila[0]: fila[1] for fila in filas}

    def _punto_por_codigo(self, codigo: str) -> PuntoVenta:
        punto = self._sesion.execute(
            select(PuntoVenta).where(PuntoVenta.codigo_co == codigo.strip())
        ).scalar_one_or_none()
        if punto is None:
            raise ErrorNoEncontrado(f"No existe el punto de venta {codigo!r}.")
        return punto

    def _categoria_por_nombre(self, nombre: str) -> Categoria:
        limpio = nombre.strip().upper()
        categoria = self._sesion.execute(
            select(Categoria).where(Categoria.nombre == limpio)
        ).scalar_one_or_none()
        if categoria is not None:
            return categoria

        destinos = destinos_de_categoria_retirada(limpio)
        if destinos:
            # No es «no existe»: es «existió y el negocio decidió retirarla».
            # Decirlo así, con el reparto delante, es la diferencia entre una
            # carga que alguien arregla en diez minutos y una que se abandona.
            raise ErrorValidacion(motivo_categoria_retirada(limpio, destinos))
        raise ErrorNoEncontrado(f"No existe la categoría {nombre!r}.")


def _clave(titulo: object) -> str:
    """Encabezado reducido a llave comparable: `'PPTO EN KILO'` → `'ppto en kilo'`."""
    return clave_columna(titulo)


def _celda(valores: Sequence[object], columnas: dict[str, int], nombre: str) -> object:
    indice = columnas.get(nombre)
    if indice is None or indice >= len(valores):
        return None
    return valores[indice]


def _texto_celda(valor: object) -> str | None:
    """Texto saneado de una celda. `402` (número) y `'402'` dan lo mismo."""
    return normalizar_texto(valor)


def _a_decimal(valor: object, campo: str) -> Decimal:
    """Convierte a `Decimal` sin pasar por `float`.

    Se convierte desde `str` a propósito: `Decimal(0.1)` arrastra el error
    binario del `float` y en un presupuesto eso es un defecto, no un detalle.
    """
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        return Decimal("0")
    texto = str(valor).strip().replace("$", "").replace(" ", "")
    # Formato colombiano: 1.234.567,89 → 1234567.89
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        numero = Decimal(texto)
    except (InvalidOperation, ValueError) as exc:
        raise ErrorValidacion(f"El campo {campo} no es un número: {valor!r}.") from exc
    if numero < 0:
        raise ErrorValidacion(f"El campo {campo} no puede ser negativo.")
    return numero
