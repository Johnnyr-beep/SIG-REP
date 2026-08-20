"""Exportación de reportes a Excel (`GET /reportes/{cualquiera}/exportar`).

Se exporta **lo mismo que se muestra**, con los mismos filtros y a partir de la
misma respuesta ya calculada. Recalcular por otro camino es cómo se llega a que
la pantalla y el archivo digan cosas distintas, que es exactamente el problema
del que viene huyendo este proyecto.

Los números salen como números —no como texto— para que quien reciba el archivo
pueda seguir trabajando con él; los porcentajes con formato de porcentaje y la
fecha de corte visible en la primera fila, siempre (§6).
"""

from __future__ import annotations

import io
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.schemas.reportes import (
    RespuestaClientes,
    RespuestaCumplimiento,
    RespuestaTablero,
    RespuestaVentaDiaria,
)

_ENCABEZADOS_INDICADORES = [
    "Presupuesto",
    "Venta",
    "Cumplimiento",
    "Ideal",
    "Brecha",
    "Semáforo",
    "Proyección",
    "Cumpl. proyectado",
    "Venta diaria promedio",
    "Venta diaria requerida",
    "Venta año anterior",
    "Crecimiento",
    "Margen valor",
    "Margen %",
    "Días hábiles",
    "Días trabajados",
]

_FORMATO_PORCENTAJE = "0.00%"
_FORMATO_MONEDA = "#,##0.00"
#: Columnas de `_ENCABEZADOS_INDICADORES` que son porcentaje (base 0).
_INDICES_PORCENTAJE = {2, 3, 4, 7, 11, 13}


def exportar_tablero(datos: RespuestaTablero) -> bytes:
    libro, hoja = _libro("Tablero", datos.periodo, str(datos.fecha_corte), datos.medida.value)

    hoja.append(["Nivel", "Código", "Nombre", *_ENCABEZADOS_INDICADORES])
    _formatear_encabezado(hoja)

    _fila_indicadores(hoja, ["Consolidado", "", "COMPAÑÍA"], datos.consolidado)
    for grupo in datos.grupos:
        _fila_indicadores(hoja, ["Grupo", grupo.codigo, grupo.nombre], grupo)

    if datos.sin_presupuesto:
        hoja.append([])
        hoja.append(["Puntos de venta sin presupuesto (venta informativa)"])
        hoja.append(["C.O.", "Nombre", "Venta", "Kilos"])
        for punto in datos.sin_presupuesto:
            hoja.append([punto.codigo_co, punto.nombre, punto.venta, punto.kilos])

    return _cerrar(libro, hoja, columnas=3 + len(_ENCABEZADOS_INDICADORES))


def exportar_cumplimiento(datos: RespuestaCumplimiento) -> bytes:
    libro, hoja = _libro("Cumplimiento", datos.periodo, str(datos.fecha_corte), datos.medida.value)

    hoja.append(["Nivel", "Punto de venta", "Nombre", *_ENCABEZADOS_INDICADORES])
    _formatear_encabezado(hoja)

    for fila in datos.filas:
        _fila_indicadores(hoja, ["PDV", fila.punto_venta, fila.nombre], fila)
        for categoria in fila.categorias:
            _fila_indicadores(hoja, ["Categoría", fila.punto_venta, categoria.categoria], categoria)

    return _cerrar(libro, hoja, columnas=3 + len(_ENCABEZADOS_INDICADORES))


def _periodos_de(datos: RespuestaVentaDiaria) -> list[str]:
    """Los períodos que toca el rango, en orden. Uno solo en el modo de siempre."""
    return list(datos.periodos) if datos.periodos else [datos.periodo]


def _referencia(datos: RespuestaVentaDiaria, codigo: str, periodo: str) -> Decimal | None:
    """Presupuesto diario de un punto en un período concreto."""
    del_periodo = datos.presupuesto_diario_por_periodo.get(periodo)
    if del_periodo is not None:
        return del_periodo.get(codigo)
    return datos.presupuesto_diario_por_pdv.get(codigo)


def exportar_venta_diaria(datos: RespuestaVentaDiaria) -> bytes:
    libro, hoja = _libro("Venta diaria", datos.periodo, str(datos.fecha_corte), datos.medida.value)

    # El presupuesto es mensual y el rango puede no serlo. Con un solo período va
    # una columna de referencia, como siempre; cuando el rango cruza de mes van
    # **todas**, una por período, porque un día de julio no se mide contra el
    # presupuesto de agosto. Escribir una sola sería publicar la referencia
    # equivocada para la mitad de las columnas, y promediarlas seria inventar un
    # número que no existe en ningún sitio. Es el mismo criterio que aplica la
    # pantalla; si el archivo exportado dijera otra cosa, quien lo abra tendria
    # dos verdades y ninguna forma de saber cual vale.
    periodos = _periodos_de(datos)
    encabezado_ppto = (
        ["Ppto. diario"]
        if len(periodos) <= 1
        else [f"Ppto. diario {periodo}" for periodo in periodos]
    )

    hoja.append(
        [
            "Punto de venta",
            "Nombre",
            *encabezado_ppto,
            *[str(f) for f in datos.fechas],
            "Total",
        ]
    )
    _formatear_encabezado(hoja)

    for fila in datos.filas:
        hoja.append(
            [
                fila.punto_venta,
                fila.nombre,
                *[_referencia(datos, fila.punto_venta, periodo) for periodo in periodos],
                *fila.valores,
                fila.total,
            ]
        )

    # La fila de totales, la misma que publica la respuesta: se exporta en lugar
    # de recalcularse aquí para que el archivo no pueda discrepar de la pantalla.
    totales = datos.totales
    ppto_totales = (
        [totales.presupuesto_diario]
        if len(periodos) <= 1
        else [totales.presupuesto_diario_por_periodo.get(periodo) for periodo in periodos]
    )
    hoja.append(
        [
            "",
            f"TOTAL · {len(datos.filas)} punto(s)",
            *ppto_totales,
            *totales.valores,
            totales.total,
        ]
    )
    for celda in hoja[hoja.max_row]:
        celda.font = Font(bold=True)

    return _cerrar(libro, hoja, columnas=3 + len(periodos) + len(datos.fechas))


def exportar_clientes(datos: RespuestaClientes) -> bytes:
    libro, hoja = _libro("Clientes", datos.periodo, str(datos.fecha_corte), datos.por)

    hoja.append(["Clave", "Nombre", "Venta", "Kilos", "Margen %", "Participación"])
    _formatear_encabezado(hoja)

    for fila in datos.filas:
        hoja.append(
            [
                fila.clave,
                fila.nombre,
                fila.venta,
                fila.kilos,
                fila.margen_porcentaje,
                fila.participacion,
            ]
        )
        for columna in (5, 6):
            hoja.cell(row=hoja.max_row, column=columna).number_format = _FORMATO_PORCENTAJE
    return _cerrar(libro, hoja, columnas=6)


# ── Interno ───────────────────────────────────────────────────────────────────


def _libro(titulo: str, periodo: str, fecha_corte: str, medida: str) -> tuple[Workbook, Any]:
    libro = Workbook()
    # `Workbook.active` está tipada como opcional porque un libro puede quedarse
    # sin hojas. El recién creado siempre trae una, pero se resuelve explícito
    # para no apoyarse en esa suposición.
    hoja = libro.active if libro.active is not None else libro.create_sheet()
    hoja.title = titulo[:31]
    # La fecha de corte va arriba y siempre: «un reporte sin fecha de corte es
    # un reporte que alguien va a malinterpretar» (§6).
    hoja.append([f"SIGREP · {titulo}"])
    hoja.append([f"Período: {periodo}", f"Fecha de corte: {fecha_corte}", f"Medida: {medida}"])
    hoja.append([])
    hoja["A1"].font = Font(bold=True, size=14)
    return libro, hoja


def _formatear_encabezado(hoja: Any) -> None:
    for celda in hoja[hoja.max_row]:
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center", wrap_text=True)


def _fila_indicadores(hoja: Any, prefijo: list[str], fila: Any) -> None:
    valores: list[Any] = [
        fila.presupuesto,
        fila.venta,
        fila.cumplimiento,
        fila.ideal,
        fila.brecha,
        fila.semaforo.value,
        fila.proyeccion,
        fila.cumplimiento_proyectado,
        fila.venta_diaria_promedio,
        fila.venta_diaria_requerida,
        fila.venta_anio_anterior,
        fila.crecimiento,
        fila.margen_valor,
        fila.margen_porcentaje,
        fila.dias_habiles,
        fila.dias_trabajados,
    ]
    hoja.append([*prefijo, *valores])

    numero_fila = hoja.max_row
    for indice, valor in enumerate(valores):
        celda = hoja.cell(row=numero_fila, column=len(prefijo) + indice + 1)
        if indice in _INDICES_PORCENTAJE:
            celda.number_format = _FORMATO_PORCENTAJE
        elif isinstance(valor, Decimal):
            celda.number_format = _FORMATO_MONEDA


def _cerrar(libro: Workbook, hoja: Any, *, columnas: int) -> bytes:
    hoja.freeze_panes = "A5"
    for indice in range(1, columnas + 1):
        hoja.column_dimensions[get_column_letter(indice)].width = 18

    memoria = io.BytesIO()
    libro.save(memoria)
    libro.close()
    return memoria.getvalue()
