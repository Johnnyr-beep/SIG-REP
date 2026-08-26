"""Lector del libro anual de agropecuaria para la importación comercial.

Módulo **puro**: abre el `.xlsx` con `openpyxl` en modo `read_only` y
`data_only`, localiza la hoja `RESUMEN (MES)` (o una equivalente robusta),
identifica la fila de encabezado con los meses `ENE..DIC` y, para un período
`YYYY-MM`, devuelve un canal por fila con el valor del mes elegido leído **tal
cual está almacenado** —sin escalar por 1 000—.

No toca la base de datos ni sabe nada de SQLAlchemy: así se puede probar
con un libro generado sin montar el esquema.

── La estructura del libro ────────────────────────────────────────────────────

La hoja `RESUMEN (MES)` trae:

- Una fila de encabezado con `DATO, ENE, FEB, MAR, Q1, ABR, ..., DIC, Q4,
  TOTAL`. Los meses `ENE..DIC` son las columnas que importan; `Q1..Q4` y
  `TOTAL` son subtotales que se ignoran.
- Cada canal ocupa **tres filas**: la primera con el nombre en la columna A y
  el año histórico (`2025`), la segunda con el año del presupuesto (`2026`) y
  la columna A vacía, y la tercera con `% Crec`. Algunos canales solo traen la
  fila del año del presupuesto.
- Los meses son columnas y los canales son filas: `SUPER MAYORISTA`,
  `MAYORISTA`, `TAT`, `Call Center`, etc. También hay filas de subtotal
  (`TOTAL DISTRIBUCIÓN`, `TOTAL TAT`…) que llegan aquí sin distinguirse: la
  importación las trata como canales y, si no están mapeadas, las rechaza con
  su motivo.

La fila del presupuesto es la cuyo segundo valor es el año del período. Así se
evita leer el histórico (`2025`) o la fila de `% Crec` (texto) como presupuesto.
"""

from __future__ import annotations

import io
import re
import unicodedata
from collections.abc import Sequence
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from app.application.services.periodos import parsear_periodo
from app.core.errors import ErrorValidacion

#: Nombres de los meses en español, en el orden de las columnas del libro. Son
#: las grafías que trae el encabezado de `RESUMEN (MES)`; la comparación es
#: normalizada (mayúsculas, sin tildes), así que `Ene` y `ENE` son la misma.
_MESES: tuple[str, ...] = (
    "ENE",
    "FEB",
    "MAR",
    "ABR",
    "MAY",
    "JUN",
    "JUL",
    "AGO",
    "SEP",
    "OCT",
    "NOV",
    "DIC",
)

#: Espacios que hay que recortar, incluidos el no separable (U+00A0) y el de
#: ancho cero (U+200B), que llegan pegados a los textos copiados desde Excel.
_ESPACIOS = " \t\r\n\xa0\u200b"
_MULTIESPACIO = re.compile(r"\s+")

CERO = Decimal("0")
ESCALA_DINERO = Decimal("0.01")


@dataclass(frozen=True, slots=True)
class CanalLeido:
    """Un canal del Excel con el valor del mes elegido, ya como `Decimal`.

    `canal` es el **nombre crudo** tal como lo trae el libro, sin normalizar: el
    que se muestra al usuario en el resultado de la importación. La
    normalización para buscar el mapeo la hace quien llama, con
    `normalizar_canal`.
    """

    canal: str
    #: Número de fila del libro, para mensajes de error accionables.
    fila: int
    monto: Decimal


def normalizar_canal(valor: object, limite: int = 120) -> str:
    """Nombre del canal a llave comparable: mayúsculas, sin tildes, espacios colapsados.

    Es la forma con la que se guarda y se busca el mapeo: `SUPER MAYORISTA` y
    `super mayorista ` (con espacio de relleno) son el mismo canal. Dos canales
    que se parecen son dos canales distintos: no hay distancia de edición ni
    prefijos, igual que en `normalizar_clave`.
    """
    if valor is None:
        return ""
    crudo = _a_texto(valor).strip(_ESPACIOS)
    if not crudo:
        return ""
    sin_tildes = unicodedata.normalize("NFKD", crudo)
    sin_tildes = "".join(c for c in sin_tildes if not unicodedata.combining(c))
    colapsado = _MULTIESPACIO.sub(" ", sin_tildes).strip().upper()
    return colapsado[:limite]


def leer_canales(contenido: bytes, codigo_periodo: str) -> list[CanalLeido]:
    """Lee el libro y devuelve un canal por fila con el valor del mes elegido.

    El valor se lee **tal cual está almacenado**: no se escala por 1 000 ni se
    interpreta. Una celda vacía se omite (ese canal no trae presupuesto para el
    mes); una celda con un número se devuelve como `Decimal` sin pasar por
    `float`.
    """
    anio, mes = parsear_periodo(codigo_periodo)

    hoja, fila_encabezado, mapa_meses = _abrir_hoja(contenido)
    nombre_mes = _MESES[mes - 1]
    if nombre_mes not in mapa_meses:
        raise ErrorValidacion(
            f"El encabezado del libro no tiene la columna del mes {nombre_mes}. Los "
            "meses admitidos son ENE, FEB, MAR, ABR, MAY, JUN, JUL, AGO, SEP, OCT, "
            "NOV y DIC."
        )
    columna_mes = mapa_meses[nombre_mes]

    canales: list[CanalLeido] = []
    canal_actual = ""
    for numero, valores in enumerate(hoja.iter_rows(values_only=True), start=1):
        if numero <= fila_encabezado:
            continue
        col_a = valores[0] if len(valores) > 0 else None
        col_b = valores[1] if len(valores) > 1 else None

        if col_a is not None and str(col_a).strip(_ESPACIOS):
            canal_actual = str(col_a).strip(_ESPACIOS)

        if canal_actual and _es_fila_del_anio(col_b, anio):
            valor = valores[columna_mes] if columna_mes < len(valores) else None
            if valor is None:
                continue
            monto = _a_decimal(valor, canal_actual, numero)
            canales.append(CanalLeido(canal=canal_actual, fila=numero, monto=monto))

    return canales


# ── Interno: apertura y localización de la hoja ──────────────────────────────


def _abrir_hoja(contenido: bytes) -> tuple[Any, int, dict[str, int]]:
    """Abre el libro y devuelve (hoja, fila_encabezado, mapa_mes→columna).

    El libro se cierra fuera de aquí: el llamador itera la hoja y la cierra al
    terminar. La hoja se devuelve abierta porque `iter_rows` la necesita viva.
    """
    from openpyxl import load_workbook

    try:
        libro = load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl lanza de todo ante un archivo roto
        raise ErrorValidacion(
            "No se pudo abrir el libro: está dañado o incompleto. Vuelva a "
            "guardarlo desde Excel y súbalo de nuevo."
        ) from exc

    try:
        hoja = _elegir_hoja(libro)
        fila_encabezado, mapa_meses = _localizar_encabezado(hoja)
        return hoja, fila_encabezado, mapa_meses
    except ErrorValidacion:
        libro.close()
        raise
    except Exception as exc:
        libro.close()
        raise ErrorValidacion(
            "No se pudo leer la hoja de presupuesto: el libro no tiene la forma "
            "esperada (encabezado con ENE..DIC y filas por canal)."
        ) from exc


def _elegir_hoja(libro: Any) -> Any:
    """Selecciona la hoja `RESUMEN (MES)` o una equivalente robusta.

    Preferencia exacta → nombre que contenga `RESUMEN` y `MES` → primera hoja.
    """
    nombres = list(libro.sheetnames)
    for nombre in nombres:
        if normalizar_canal(nombre) == "RESUMEN (MES)":
            return libro[nombre]
    for nombre in nombres:
        clave = normalizar_canal(nombre)
        if "RESUMEN" in clave and "MES" in clave:
            return libro[nombre]
    if not nombres:
        raise ErrorValidacion("El libro no tiene hojas.")
    return libro[nombres[0]]


def _localizar_encabezado(hoja: Any) -> tuple[int, dict[str, int]]:
    """Encuentra la fila con `ENE..DIC` y devuelve (fila, mapa mes→columna).

    Recorre las primeras 20 filas buscando la que tenga al menos `ENE` y `DIC`;
    de ella saca el mapa mes→columna para resolver luego la columna del período.
    """
    for numero, valores in enumerate(hoja.iter_rows(values_only=True), start=1):
        if numero > 20:
            break
        mapa = _mapa_meses(valores)
        if "ENE" in mapa and "DIC" in mapa:
            return numero, mapa
    raise ErrorValidacion(
        "No se encontró la fila de encabezado con los meses ENE..DIC. El libro "
        "debe tener una hoja con los meses como columnas."
    )


def _mapa_meses(valores: Sequence[Any]) -> dict[str, int]:
    """`{mes normalizado: indice de columna}` para una fila de encabezado."""
    mapa: dict[str, int] = {}
    for indice, celda in enumerate(valores):
        clave = normalizar_canal(celda, limite=10)
        if clave in _MESES and clave not in mapa:
            mapa[clave] = indice
    return mapa


# ── Interno: conversión de valores ───────────────────────────────────────────


def _es_fila_del_anio(col_b: object, anio: int) -> bool:
    """¿Esta fila es la del año del período?

    `col_b` llega como entero (`2026`), flotante (`2026.0`) o texto (`'2026'`).
    Los tres son el mismo año. La fila de `% Crec` llega como texto y se
    descarta aquí, no en el llamador: es la que distingue el presupuesto del
    histórico y del crecimiento.
    """
    if isinstance(col_b, bool):
        return False
    if isinstance(col_b, int | float):
        return int(col_b) == anio
    crudo = normalizar_canal(col_b, limite=10).replace(" ", "")
    if crudo.endswith(".0"):
        crudo = crudo[:-2]
    return crudo.isdigit() and int(crudo) == anio


def _a_decimal(valor: object, canal: str, fila: int) -> Decimal:
    """Valor de la celda a `Decimal` sin pasar por `float` y sin escalar.

    **No se multiplica por 1 000**: el valor se lee tal cual está almacenado. El
    libro del negocio trae los importes ya en pesos, no en miles de pesos.
    """
    if isinstance(valor, Decimal):
        numero = valor
    elif isinstance(valor, int):
        numero = Decimal(valor)
    elif isinstance(valor, float):
        # `float` no promete exactitud, pero es lo que entrega openpyxl para una
        # celda numérica. Se reconstruye desde `repr` para no arrastrar el error
        # binario más de lo inevitable.
        numero = Decimal(repr(valor))
    else:
        texto = str(valor).strip().replace("$", "").replace(" ", "")
        if "," in texto and "." in texto:
            texto = texto.replace(".", "").replace(",", ".")
        elif "," in texto:
            texto = texto.replace(",", ".")
        try:
            numero = Decimal(texto)
        except (InvalidOperation, ValueError) as exc:
            raise ErrorValidacion(
                f"El valor del canal «{canal}» en la fila {fila} no es un número: {valor!r}."
            ) from exc
    if numero < CERO:
        raise ErrorValidacion(
            f"El valor del canal «{canal}» en la fila {fila} no puede ser negativo: {valor!r}."
        )
    return numero.quantize(ESCALA_DINERO)


def _a_texto(valor: object) -> str:
    """`2026.0` → `'2026'`; el resto, `str()` a secas."""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor)
