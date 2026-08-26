"""Presupuesto agropecuario: cuatro descomposiciones del mismo total.

Este módulo es donde se hace cumplir la regla que da forma a toda la unidad:

    **El presupuesto por vendedor, por centro de operación, por especie y por
    tipo comercial describen EL MISMO DINERO. No se suman entre sí.**

── Cómo se hace imposible, y no solo se advierte ─────────────────────────────

Un comentario que dijera «ojo, no sumes esto» duraría hasta el primer lunes con
prisa. Lo que hay en su lugar son cuatro cierres, y hay que saltárselos todos
para cometer el error:

1. **No existe ninguna consulta sin dimensión.** `_filas` es el único sitio del
   módulo que hace `SELECT ... FROM agro_presupuestos`, y su firma exige un
   `DimensionPresupuesto`. No hay un `listar_todo()` que devuelva importes
   sueltos; la lectura pública (`listar`) devuelve un mapa **por dimensión** con
   un total dentro de cada una y ninguno global.

2. **El presupuesto no circula como número, circula como `PlanPresupuesto`**, un
   objeto inmutable que lleva su dimensión pegada. Todo lo que el reporte
   necesita —el total, la meta de un miembro— sale de él.

3. **`PlanPresupuesto` se niega a sumarse.** `plan_a + plan_b` lanza
   `ErrorDimensionesIncompatibles`, y `sum(planes)` también, porque `__radd__`
   —el que recibe el `0` inicial de `sum`— lanza igual. No hay ninguna ruta por
   la que dos planes acaben convertidos en un número mayor que cualquiera de
   los dos.

4. **Y cuando las cuatro descomposiciones no dan lo mismo, se dice.** `cuadre`
   compara los totales de las dimensiones capturadas y publica la diferencia con
   nombre y apellido. **No la corrige**: un descuadre es un error de captura del
   negocio, y repartirlo por cuenta del sistema sería inventarse la meta de
   alguien. Se hace visible —también dentro de `parametros_calculo` de cada
   reporte— y quien lo capturó lo arregla.

── Lo que sí se reutiliza de carnes ──────────────────────────────────────────

El versionado con autor y motivo (§3.3), el bloqueo por período cerrado (§7) y
la carga masiva. Son las mismas reglas, no parecidas, y aquí se aplican sobre la
tabla agropecuaria con el mismo criterio: solo se historia lo que **cambió de
valor**, el motivo es obligatorio y no admite una palabra suelta, y una fila
mala de la carga masiva no aborta las buenas.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.application.services import agro_presupuesto_tabla as tabla
from app.application.services.periodos import obtener_o_crear_periodo, obtener_periodo
from app.core.errors import (
    ErrorNoEncontrado,
    ErrorPeriodoCerrado,
    ErrorSigrep,
    ErrorValidacion,
)
from app.core.logging import obtener_logger
from app.domain.normalizacion import (
    ESCALA_DINERO,
    ESCALA_KILOS,
    clave_columna,
    normalizar_texto,
)
from app.infrastructure.models.agro_dimensiones import AgroDimension
from app.infrastructure.models.agro_presupuesto import AgroPresupuesto, AgroPresupuestoHistorial
from app.infrastructure.models.agro_vocabulario import (
    DimensionPresupuesto,
    TipoDimension,
    normalizar_clave,
    normalizar_etiqueta,
)
from app.infrastructure.models.mixins import ahora_utc
from app.infrastructure.models.periodo import Periodo
from app.infrastructure.models.usuario import Usuario
from app.schemas.agro import (
    CuadreDimension,
    CuadrePresupuestoSalida,
    ErrorFilaAgro,
    HistorialAgroSalida,
    MiembroDimensionSalida,
    PresupuestoAgroSalida,
    PresupuestoDimensionSalida,
    ResultadoCargaAgro,
)

logger = obtener_logger(__name__)

CERO = Decimal("0")

#: Encabezados admitidos en la carga masiva. Se aceptan varias grafías porque el
#: archivo lo arma el negocio a mano y exigirle una plantilla exacta es
#: garantizar que la primera carga falle.
_ALIAS_COLUMNAS: dict[str, str] = {
    "dimension": "dimension",
    "dimensión": "dimension",
    "eje": "dimension",
    "clave": "clave",
    "codigo": "clave",
    "código": "clave",
    "id": "clave",
    "etiqueta": "etiqueta",
    "nombre": "etiqueta",
    "descripcion": "etiqueta",
    "descripción": "etiqueta",
    "monto": "monto",
    "ppto": "monto",
    "presupuesto": "monto",
    "valor": "monto",
    "kilos": "kilos",
    "ppto en kilo": "kilos",
    "ppto kilos": "kilos",
}

#: Grafías con las que el negocio nombra cada dimensión en su archivo. La carga
#: masiva **no adivina**: un valor que no esté aquí rechaza la fila diciendo
#: cuáles son los cuatro admitidos, porque una dimensión mal leída pondría la
#: meta de un vendedor en el renglón de una especie.
_ALIAS_DIMENSION: dict[str, DimensionPresupuesto] = {
    "vendedor": DimensionPresupuesto.VENDEDOR,
    "vendedores": DimensionPresupuesto.VENDEDOR,
    "centro_operacion": DimensionPresupuesto.CENTRO_OPERACION,
    "centro de operacion": DimensionPresupuesto.CENTRO_OPERACION,
    "centro de operación": DimensionPresupuesto.CENTRO_OPERACION,
    "centro": DimensionPresupuesto.CENTRO_OPERACION,
    "co": DimensionPresupuesto.CENTRO_OPERACION,
    "especie": DimensionPresupuesto.ESPECIE,
    "especies": DimensionPresupuesto.ESPECIE,
    "tipo_comercial": DimensionPresupuesto.TIPO_COMERCIAL,
    "tipo comercial": DimensionPresupuesto.TIPO_COMERCIAL,
}


class ErrorDimensionesIncompatibles(ErrorValidacion):
    """Se intentó operar aritméticamente con presupuestos de distinta dimensión.

    Tiene código propio —`dimensiones_incompatibles`— y no el genérico de
    validación porque no es un dato mal escrito: es la operación que este módulo
    existe para impedir. Que llegue hasta una respuesta HTTP significa que
    alguien encontró una ruta nueva hacia el error, y quien la lea tiene que
    poder distinguirla de un campo mal tecleado.
    """

    codigo = "dimensiones_incompatibles"


@dataclass(frozen=True, slots=True)
class MetaMiembro:
    """La meta de un miembro dentro de **una** dimensión."""

    clave: str
    etiqueta: str | None
    monto: Decimal
    kilos: Decimal


@dataclass(frozen=True, slots=True)
class PlanPresupuesto:
    """El presupuesto de un período **visto por una sola dimensión**.

    Es inmutable y lleva su dimensión pegada. Esa es toda la idea: un importe de
    presupuesto no circula por el sistema como un `Decimal` suelto —que se
    sumaría con cualquier otro sin que nada chirriara— sino dentro de un objeto
    que sabe de qué descomposición es y se niega a mezclarse con otra.

    `total` es el presupuesto **de la compañía**, no de un trozo: las cuatro
    dimensiones reparten el mismo total, así que el total de cualquiera de ellas
    ya es el de la compañía. Por eso sumar dos planes no es «juntar dos partes»,
    es contar el mismo dinero dos veces.
    """

    dimension: DimensionPresupuesto
    metas: Mapping[str, MetaMiembro]

    @property
    def definido(self) -> bool:
        """¿Hay alguna meta capturada en esta dimensión?

        Distingue «el negocio presupuestó cero» de «nadie ha presupuestado
        todavía». En el segundo caso el cumplimiento viaja vacío —«—»— y el
        semáforo sale `SIN_PRESUPUESTO`, nunca en rojo: no hay vara.
        """
        return bool(self.metas)

    def de(self, clave: str) -> MetaMiembro | None:
        """La meta de un miembro, o `None` si ese miembro no está presupuestado."""
        return self.metas.get(clave)

    def monto_de(self, clave: str) -> Decimal | None:
        meta = self.metas.get(clave)
        return meta.monto if meta is not None else None

    def kilos_de(self, clave: str) -> Decimal | None:
        meta = self.metas.get(clave)
        return meta.kilos if meta is not None else None

    @property
    def total_monto(self) -> Decimal:
        return sum((meta.monto for meta in self.metas.values()), start=CERO)

    @property
    def total_kilos(self) -> Decimal:
        return sum((meta.kilos for meta in self.metas.values()), start=CERO)

    # ── La suma prohibida ─────────────────────────────────────────────────────

    def __add__(self, otro: object) -> PlanPresupuesto:
        """Nunca. Ni con otra dimensión ni con la misma.

        Con **otra dimensión** es el error que este módulo existe para impedir:
        el presupuesto por vendedor y el presupuesto por especie son el mismo
        dinero visto de dos formas y sumarlos da el doble de la meta real.

        Con **la misma dimensión** también se niega, y es deliberado: dos planes
        de la misma dimensión solo pueden ser de dos períodos distintos, y
        «meta de enero + meta de febrero» es una operación legítima que no
        necesita un operador ambiguo. Quien la quiera, que sume `total_monto`,
        que dice exactamente lo que hace.
        """
        raise ErrorDimensionesIncompatibles(_motivo_suma(self, otro))

    def __radd__(self, otro: object) -> PlanPresupuesto:
        """El `0` inicial de `sum(...)` entra por aquí, y también se niega.

        Sin este método, `sum(planes)` haría `0 + plan` —que Python resolvería
        como `int.__add__`, `NotImplemented`, y luego este `__radd__`— y sin él
        acabaría en un `TypeError` opaco. Con él, el mensaje dice qué se
        intentó y por qué no se puede.
        """
        raise ErrorDimensionesIncompatibles(_motivo_suma(self, otro))


def _motivo_suma(plan: PlanPresupuesto, otro: object) -> str:
    """El mensaje de la suma prohibida, que tiene que enseñar la salida."""
    if isinstance(otro, PlanPresupuesto) and otro.dimension is not plan.dimension:
        return (
            f"No se puede sumar el presupuesto por {plan.dimension.etiqueta.lower()} con el "
            f"presupuesto por {otro.dimension.etiqueta.lower()}: son la misma meta de la "
            "compañía repartida de dos formas distintas, y sumarlas daría el doble. El "
            "cumplimiento se calcula **dentro** de una dimensión, comparando la venta "
            "agregada por esa misma dimensión contra su presupuesto. Si lo que busca es "
            "comprobar que las dos descomposiciones dan el mismo total, use "
            "`GET /agro/presupuesto/cuadre`."
        )
    if isinstance(otro, PlanPresupuesto):
        return (
            f"No se suman dos planes de presupuesto, ni siquiera de la misma dimensión "
            f"({plan.dimension.etiqueta.lower()}). Dos planes de la misma dimensión son de "
            "dos períodos distintos, y el acumulado de varios meses se pide sumando "
            "`total_monto`, que dice lo que hace, en lugar de un `+` que se lee igual que "
            "la suma prohibida entre dimensiones."
        )
    return (
        "Un `PlanPresupuesto` no es un número y no participa en sumas. Lleva su dimensión "
        "pegada justamente para que el presupuesto de una descomposición no acabe sumado "
        "con el de otra: son el mismo dinero visto de dos formas. Use `total_monto` o "
        "`total_kilos` si lo que necesita es el importe."
    )


@dataclass(frozen=True, slots=True)
class Cuadre:
    """Resultado de comparar las descomposiciones capturadas entre sí.

    `cuadra` es `True` cuando **todas** las dimensiones con presupuesto
    capturado dan el mismo total, en pesos y en kilos. Con una sola dimensión
    capturada cuadra por definición: no hay con qué contrastar, y decir que no
    cuadra sería llamar error a un presupuesto a medio capturar.

    La comparación es **exacta**. No hay tolerancia y no debe haberla: estos son
    importes que alguien tecleó a partir de un reparto que él mismo hizo, y una
    diferencia de un peso es tan error de captura como una de seiscientos
    millones. Lo que sí se publica es **cuánta** es la diferencia, para que se
    distinga de un vistazo un dedazo de un renglón que falta.
    """

    periodo: str
    totales: Mapping[DimensionPresupuesto, tuple[Decimal, Decimal]]
    cuadra: bool
    diferencia_monto: Decimal
    diferencia_kilos: Decimal
    mensaje: str


class AgroPresupuestoService:
    """Casos de uso del presupuesto agropecuario."""

    def __init__(self, sesion: Session) -> None:
        self._sesion = sesion

    # ── Lectura para el reporte ───────────────────────────────────────────────

    def plan(self, periodo: Periodo, dimension: DimensionPresupuesto) -> PlanPresupuesto:
        """El presupuesto del período **visto por una dimensión**.

        Es la única forma de sacar presupuesto de esta tabla, y `dimension` es
        obligatoria y tipada: no hay manera de pedir «todo el presupuesto» y
        acabar con importes de dos descomposiciones en la misma bolsa.
        """
        metas = {
            fila.clave: MetaMiembro(
                clave=fila.clave,
                etiqueta=fila.etiqueta,
                monto=Decimal(fila.monto),
                kilos=Decimal(fila.kilos),
            )
            for fila in self._filas(periodo, dimension)
        }
        return PlanPresupuesto(dimension=dimension, metas=metas)

    def _filas(
        self, periodo: Periodo, dimension: DimensionPresupuesto
    ) -> Sequence[AgroPresupuesto]:
        """**El único `SELECT` sobre `agro_presupuestos` de todo el módulo.**

        Y siempre con `dimension` en el `WHERE`. Que solo exista este método es
        lo que garantiza que ninguna consulta agregue presupuestos de dos
        descomposiciones: no hay otra puerta.
        """
        return list(
            self._sesion.execute(
                select(AgroPresupuesto)
                .where(
                    AgroPresupuesto.periodo_id == periodo.id,
                    AgroPresupuesto.dimension == dimension.value,
                )
                .order_by(AgroPresupuesto.clave)
            ).scalars()
        )

    # ── Lectura para la pantalla ──────────────────────────────────────────────

    def listar(
        self, codigo_periodo: str, dimension: DimensionPresupuesto | None = None
    ) -> list[PresupuestoDimensionSalida]:
        """Presupuesto del período, **agrupado por dimensión**.

        Devuelve una entrada por dimensión, cada una con **su** total. No hay un
        total global y no puede haberlo: sería la suma prohibida. Una pantalla
        que quiera «el presupuesto de la compañía» elige una dimensión, que es
        la operación correcta —las cuatro reparten el mismo total—.
        """
        periodo = obtener_periodo(self._sesion, codigo_periodo)
        dimensiones = [dimension] if dimension is not None else list(DimensionPresupuesto)
        etiquetas = self._etiquetas_catalogo()

        salida: list[PresupuestoDimensionSalida] = []
        for actual in dimensiones:
            plan = self.plan(periodo, actual)
            filas = [
                PresupuestoAgroSalida(
                    dimension=actual.value,
                    clave=meta.clave,
                    nombre=self._nombre(actual, meta, etiquetas),
                    monto=meta.monto,
                    kilos=meta.kilos,
                )
                for meta in sorted(plan.metas.values(), key=lambda m: m.clave)
            ]
            salida.append(
                PresupuestoDimensionSalida(
                    dimension=actual.value,
                    etiqueta=actual.etiqueta,
                    definido=plan.definido,
                    total_monto=plan.total_monto,
                    total_kilos=plan.total_kilos,
                    filas=filas,
                )
            )
        return salida

    def historial(
        self,
        codigo_periodo: str | None = None,
        dimension: DimensionPresupuesto | None = None,
        clave: str | None = None,
    ) -> list[HistorialAgroSalida]:
        """Rastro de cambios: quién, cuándo, qué campo y por qué (§3.3)."""
        consulta = select(AgroPresupuestoHistorial).order_by(AgroPresupuestoHistorial.cuando.desc())
        if codigo_periodo:
            periodo = obtener_periodo(self._sesion, codigo_periodo)
            consulta = consulta.where(AgroPresupuestoHistorial.periodo_id == periodo.id)
        if dimension is not None:
            consulta = consulta.where(AgroPresupuestoHistorial.dimension == dimension.value)
        if clave:
            # Se normaliza con la misma función con la que se guardó, y con el
            # tipo de la dimensión cuando se conoce: el C.O. se rellena a tres
            # posiciones y sin ese detalle `301` no encontraría a `'301'`.
            buscada = (
                normalizar_clave(dimension.tipo, clave)
                if dimension is not None
                else (normalizar_etiqueta(clave, limite=60) or "").upper()
            )
            consulta = consulta.where(AgroPresupuestoHistorial.clave == buscada)

        autores = self._autores()
        return [
            HistorialAgroSalida(
                cuando=h.cuando,
                quien=autores.get(h.usuario_id),
                dimension=h.dimension,
                clave=h.clave,
                campo=h.campo,
                valor_anterior=h.valor_anterior,
                valor_nuevo=h.valor_nuevo,
                motivo=h.motivo,
            )
            for h in self._sesion.execute(consulta).scalars()
        ]

    # ── El cuadre entre dimensiones ───────────────────────────────────────────

    def cuadre(self, codigo_periodo: str) -> Cuadre:
        """Compara los totales de las descomposiciones capturadas.

        Es la comprobación que el negocio va a agradecer: si el total por
        vendedor no coincide con el total por especie, **eso es un error de
        captura**, y el sistema lo dice en lugar de arreglarlo por su cuenta.
        Repartir la diferencia sería inventarse la meta de alguien; publicarla
        es lo que hace que quien la capturó la corrija.

        Solo entran las dimensiones **con presupuesto capturado**. Una dimensión
        vacía no descuadra nada: significa que todavía no se ha parametrizado,
        que es distinto de estar mal.
        """
        periodo = obtener_periodo(self._sesion, codigo_periodo)
        totales: dict[DimensionPresupuesto, tuple[Decimal, Decimal]] = {}
        for dimension in DimensionPresupuesto:
            plan = self.plan(periodo, dimension)
            if plan.definido:
                totales[dimension] = (plan.total_monto, plan.total_kilos)

        if len(totales) < 2:
            return Cuadre(
                periodo=periodo.codigo,
                totales=totales,
                cuadra=True,
                diferencia_monto=CERO,
                diferencia_kilos=CERO,
                mensaje=_MENSAJE_CUADRE_TRIVIAL if totales else _MENSAJE_SIN_PRESUPUESTO,
            )

        montos = [monto for monto, _ in totales.values()]
        kilos = [kilo for _, kilo in totales.values()]
        diferencia_monto = max(montos) - min(montos)
        diferencia_kilos = max(kilos) - min(kilos)
        cuadra = diferencia_monto == CERO and diferencia_kilos == CERO

        return Cuadre(
            periodo=periodo.codigo,
            totales=totales,
            cuadra=cuadra,
            diferencia_monto=diferencia_monto,
            diferencia_kilos=diferencia_kilos,
            mensaje=(
                _MENSAJE_CUADRA
                if cuadra
                else _mensaje_descuadre(totales, diferencia_monto, diferencia_kilos)
            ),
        )

    def cuadre_salida(self, codigo_periodo: str) -> CuadrePresupuestoSalida:
        """El cuadre en la forma que publica la API."""
        resultado = self.cuadre(codigo_periodo)
        return CuadrePresupuestoSalida(
            periodo=resultado.periodo,
            cuadra=resultado.cuadra,
            diferencia_monto=resultado.diferencia_monto,
            diferencia_kilos=resultado.diferencia_kilos,
            mensaje=resultado.mensaje,
            dimensiones=[
                CuadreDimension(
                    dimension=dimension.value,
                    etiqueta=dimension.etiqueta,
                    total_monto=monto,
                    total_kilos=kilos,
                )
                for dimension, (monto, kilos) in sorted(
                    resultado.totales.items(), key=lambda par: par[0].value
                )
            ],
        )

    # ── Parametrización ───────────────────────────────────────────────────────

    def guardar(
        self,
        *,
        codigo_periodo: str,
        dimension: DimensionPresupuesto,
        clave: str,
        monto: Decimal,
        kilos: Decimal,
        motivo: str,
        etiqueta: str | None = None,
        usuario: Usuario | None = None,
    ) -> PresupuestoAgroSalida:
        """Fija la meta de un miembro y deja el rastro del cambio.

        Solo se historian los campos que **cambiaron de valor**: guardar sin
        modificar nada no ensucia el historial, que es lo que hace que el
        historial siga siendo legible al cabo de un año.

        La `clave` se normaliza con la misma función que usa la ingesta, y ese
        detalle es el que hace que la meta cruce con la venta: si el presupuesto
        guardara `'01'` y la venta `'1'`, el cumplimiento de ese vendedor saldría
        en cero para siempre sin que nada fallara.
        """
        periodo = obtener_o_crear_periodo(self._sesion, codigo_periodo)
        self._exigir_periodo_abierto(periodo)

        clave_normalizada = normalizar_clave(dimension.tipo, clave)
        if clave_normalizada == "SIN_DATO":
            raise ErrorValidacion(
                f"No se puede presupuestar un miembro sin identificador en la dimensión "
                f"{dimension.etiqueta.lower()}. Indique la clave del origen "
                "(`CodigoVendedor`, `CO_Id`, `Especie_Id` o `TipoComercial_Id`)."
            )

        fila = self._sesion.execute(
            select(AgroPresupuesto).where(
                AgroPresupuesto.periodo_id == periodo.id,
                AgroPresupuesto.dimension == dimension.value,
                AgroPresupuesto.clave == clave_normalizada,
            )
        ).scalar_one_or_none()

        anterior_monto = Decimal(fila.monto) if fila is not None else None
        anterior_kilos = Decimal(fila.kilos) if fila is not None else None

        if fila is None:
            fila = AgroPresupuesto(
                periodo_id=periodo.id,
                dimension=dimension.value,
                clave=clave_normalizada,
                monto=monto,
                kilos=kilos,
            )
            self._sesion.add(fila)
        else:
            fila.monto = monto
            fila.kilos = kilos

        nueva_etiqueta = normalizar_etiqueta(etiqueta)
        if nueva_etiqueta is not None:
            fila.etiqueta = nueva_etiqueta
        fila.actualizado_por_id = usuario.id if usuario else None
        self._sesion.flush()

        self._historiar(fila, "monto", anterior_monto, monto, motivo, usuario)
        self._historiar(fila, "kilos", anterior_kilos, kilos, motivo, usuario)
        self._sesion.flush()

        logger.info(
            "agro_presupuesto_guardado",
            periodo=periodo.codigo,
            dimension=dimension.value,
            clave=clave_normalizada,
        )
        return PresupuestoAgroSalida(
            dimension=dimension.value,
            clave=clave_normalizada,
            nombre=fila.etiqueta or self._nombre_catalogo(dimension, clave_normalizada),
            monto=Decimal(fila.monto),
            kilos=Decimal(fila.kilos),
        )

    def eliminar(
        self,
        *,
        codigo_periodo: str,
        dimension: DimensionPresupuesto,
        clave: str,
        motivo: str,
        usuario: Usuario | None,
    ) -> None:
        """Retira una meta activa sin destruir su rastro de auditoría."""
        periodo = obtener_periodo(self._sesion, codigo_periodo)
        self._exigir_periodo_abierto(periodo)
        clave_normalizada = normalizar_clave(dimension.tipo, clave)
        fila = self._sesion.execute(
            select(AgroPresupuesto).where(
                AgroPresupuesto.periodo_id == periodo.id,
                AgroPresupuesto.dimension == dimension.value,
                AgroPresupuesto.clave == clave_normalizada,
            )
        ).scalar_one_or_none()
        if fila is None:
            raise ErrorNoEncontrado("La meta que intenta eliminar ya no existe.")

        self._historiar(fila, "monto", Decimal(fila.monto), None, motivo, usuario)
        self._historiar(fila, "kilos", Decimal(fila.kilos), None, motivo, usuario)
        self._sesion.flush()
        for evento in self._sesion.execute(
            select(AgroPresupuestoHistorial).where(
                AgroPresupuestoHistorial.presupuesto_id == fila.id
            )
        ).scalars():
            evento.presupuesto_id = None
        self._sesion.delete(fila)
        self._sesion.flush()

    # ── Catálogo de dimensiones ───────────────────────────────────────────────

    def miembros(self, tipo: TipoDimension) -> list[MiembroDimensionSalida]:
        """El catálogo de una dimensión, en orden alfabético.

        Sale de lo que dejó la ingesta y no de una tabla que alguien mantenga a
        mano: si aquí se pudieran inventar miembros, se crearía un «Planta» que
        no case con el `301` que manda el origen y el presupuesto quedaría
        colgado de una clave que ninguna venta usa.

        Se publica el `activo` de cada fila y no un `True` fijo: el catálogo
        tiene esa columna para poder retirar un miembro sin borrarlo —borrarlo se
        llevaría por delante el presupuesto que cuelga de su clave—, y anunciar
        a todos como activos convertiría esa marca en decoración. La lista sale
        entera y con su estado; que la pantalla ofrezca o no un miembro retirado
        es decisión suya, pero con el dato delante.
        """
        filas = self._sesion.execute(
            select(AgroDimension)
            .where(AgroDimension.tipo == tipo.value)
            .order_by(AgroDimension.nombre)
        ).scalars()
        return [
            MiembroDimensionSalida(tipo=d.tipo, clave=d.clave, nombre=d.nombre, activo=d.activo)
            for d in filas
        ]

    # ── Carga masiva ──────────────────────────────────────────────────────────

    def cargar_masivo(
        self,
        contenido: bytes,
        nombre_archivo: str,
        *,
        codigo_periodo: str,
        motivo: str,
        usuario: Usuario | None = None,
    ) -> ResultadoCargaAgro:
        """Carga un `.csv` o `.xlsx` con el presupuesto del período.

        El archivo trae una columna `dimension` por fila, de modo que las cuatro
        descomposiciones caben en un solo archivo **sin mezclarse**: cada fila
        dice a cuál pertenece y se guarda en la suya. Ese es justamente el
        formato que evita el error de capturar en una hoja «el presupuesto» sin
        decir de qué reparto es.

        Una fila mala **no aborta la carga**: se acepta lo que se pueda y se
        devuelve el detalle de lo rechazado con su número de fila. Abortar entero
        por un typo en la fila 90 obligaría a repetir 89 filas buenas.

        Al terminar se ejecuta el **cuadre** y su resultado viaja en la
        respuesta. Es el momento en que hay que verlo: quien acaba de subir las
        cuatro descomposiciones es quien puede corregir la que se desvió.
        """
        periodo = obtener_o_crear_periodo(self._sesion, codigo_periodo)
        self._exigir_periodo_abierto(periodo)

        filas, errores = self._leer_archivo(contenido, nombre_archivo)

        aceptadas = 0
        for numero, fila in filas:
            try:
                self.guardar(
                    codigo_periodo=codigo_periodo,
                    dimension=fila.dimension,
                    clave=fila.clave,
                    monto=fila.monto,
                    kilos=fila.kilos,
                    motivo=motivo,
                    etiqueta=fila.etiqueta,
                    usuario=usuario,
                )
            except (ErrorNoEncontrado, ErrorValidacion) as exc:
                errores.append(ErrorFilaAgro(fila=numero, motivo=exc.mensaje))
            else:
                aceptadas += 1

        logger.info(
            "agro_carga_masiva",
            periodo=codigo_periodo,
            aceptadas=aceptadas,
            rechazadas=len(errores),
        )
        return ResultadoCargaAgro(
            aceptadas=aceptadas,
            rechazadas=len(errores),
            errores=errores,
            cuadre=self.cuadre_salida(codigo_periodo),
        )

    def _leer_archivo(
        self, contenido: bytes, nombre_archivo: str
    ) -> tuple[list[tuple[int, _FilaCarga]], list[ErrorFilaAgro]]:
        nombre = nombre_archivo.lower()
        if nombre.endswith(".csv"):
            crudas = _filas_csv(contenido)
        elif nombre.endswith((".xlsx", ".xlsm")):
            # Los dos intentos de abrir el libro van bajo la misma guarda. Un
            # `.xlsx` truncado a medio subir tiene las partes en su sitio —así
            # que pasa la comprobación de `leer_subida`— y revienta al parsearlo:
            # `openpyxl` no promete un tipo de excepción concreto y lo que llegue
            # sin envolver sale por la API como un 500 en lugar de decirle a
            # quien subió el archivo que vuelva a subirlo.
            try:
                # Antes del formato genérico se prueba el libro que arma el
                # negocio, con su tabla dinámica y sus subtotales. Pedirle que lo
                # reformatee para poder cargarlo es la petición que acaba en
                # «entonces lo sigo llevando en Excel».
                propio = self._leer_libro_del_negocio(contenido)
                if propio is not None:
                    return propio
                crudas = _filas_excel(contenido)
            except ErrorSigrep:
                # Los errores del dominio ya dicen lo suyo —«todavía no hay
                # vendedores en el catálogo»— y taparlos con el mensaje genérico
                # mandaría a buscar un archivo corrupto que está perfecto.
                raise
            except Exception as exc:
                raise ErrorValidacion(
                    "No se pudo abrir el libro: está dañado o incompleto. Vuelva a "
                    "guardarlo desde Excel y súbalo de nuevo."
                ) from exc
        else:
            raise ErrorValidacion("Formato no admitido. Use .xlsx o .csv.")

        filas: list[tuple[int, _FilaCarga]] = []
        errores: list[ErrorFilaAgro] = []
        for numero, cruda in crudas:
            try:
                filas.append((numero, _normalizar_fila(cruda)))
            except ErrorValidacion as exc:
                errores.append(ErrorFilaAgro(fila=numero, motivo=exc.mensaje))
        return filas, errores

    def _claves_de_vendedor(self) -> dict[str, str]:
        """Nombre normalizado a clave del origen, desde el catálogo de la ingesta.

        Es la traducción sin la cual la carga no sirve de nada: el libro nombra
        al vendedor y la venta lo identifica por su cédula. Guardar la meta bajo
        el nombre la dejaría colgada de una clave que ninguna venta usa, y el
        cumplimiento de esa persona saldría **cero para siempre** sin que nada
        fallara.
        """
        filas = self._sesion.execute(
            select(AgroDimension.nombre, AgroDimension.clave).where(
                AgroDimension.tipo == TipoDimension.VENDEDOR.value
            )
        )
        return {tabla.normalizar_nombre(nombre): clave for nombre, clave in filas}

    def _leer_libro_del_negocio(
        self, contenido: bytes
    ) -> tuple[list[tuple[int, _FilaCarga]], list[ErrorFilaAgro]] | None:
        """El libro con tabla dinámica, si es que lo es. `None` si no."""
        claves = self._claves_de_vendedor()
        lectura = tabla.leer(contenido, claves)
        if lectura is None:
            return None

        if not claves:
            raise ErrorValidacion(
                "Todavía no hay vendedores en el catálogo, así que no se puede "
                "saber a quién pertenece cada meta del archivo. Ejecute primero "
                "una ingesta: el catálogo lo trae el origen, no este archivo."
            )

        filas = [
            (
                meta.fila,
                _FilaCarga(
                    dimension=DimensionPresupuesto.VENDEDOR,
                    clave=meta.clave,
                    etiqueta=meta.nombre,
                    monto=meta.monto,
                    kilos=meta.kilos,
                ),
            )
            for meta in lectura.metas
        ]

        errores = [
            ErrorFilaAgro(
                fila=0,
                motivo=(
                    f"«{nombre}» no corresponde a ningún vendedor del catálogo ni cuadra "
                    "como subtotal de su grupo. Su presupuesto NO se cargó."
                ),
            )
            for nombre in lectura.sin_resolver
        ]

        # El propio libro trae su total. Contrastarlo contra lo cargado es lo que
        # detecta que algo se quedó fuera sin que nadie lo note: la carga puede
        # terminar «bien» y publicar una meta más baja que la del negocio.
        errores.extend(self._avisos_de_cuadre(lectura, filas))
        return filas, errores

    @staticmethod
    def _avisos_de_cuadre(
        lectura: tabla.LecturaTabla, filas: list[tuple[int, _FilaCarga]]
    ) -> list[ErrorFilaAgro]:
        """Avisa si lo cargado no llega al total que el propio libro declara.

        ── Por qué hay tolerancia aquí y no la hay en el cuadre entre dimensiones

        Parecen la misma comparación y no lo son. El cuadre entre dimensiones
        contrasta cuatro cifras que **alguien tecleó**, y ahí un peso de
        diferencia es un dedazo que hay que ver. Esto contrasta el total del
        libro contra la suma de sus propias filas, y las filas se redondean al
        leerlas: el libro trae `6.315.016.727,261173` por vendedor y el sistema
        guarda dos decimales, así que veintiún redondeos de hasta medio centavo
        cada uno separan la suma del total sin que falte nadie.

        La tolerancia es exactamente eso —`filas × medio paso`—, el máximo que el
        redondeo por sí solo puede explicar. Cualquier diferencia mayor no la
        produce el redondeo: falta alguien, y se dice.

        Comparar exacto convertía este aviso en ruido de cada carga, y un aviso
        que sale siempre es un aviso que nadie lee el día que es cierto.
        """
        avisos: list[ErrorFilaAgro] = []
        pares = (
            (
                "monto",
                lectura.total_libro_monto,
                sum((f.monto for _, f in filas), Decimal(0)),
                tabla.DECIMALES_MONTO,
            ),
            (
                "kilos",
                lectura.total_libro_kilos,
                sum((f.kilos for _, f in filas), Decimal(0)),
                tabla.DECIMALES_KILOS,
            ),
        )
        for etiqueta, del_libro, cargado, paso in pares:
            if del_libro is None:
                continue
            tolerancia = paso * len(filas) / 2
            if abs(del_libro - cargado) > tolerancia:
                avisos.append(
                    ErrorFilaAgro(
                        fila=0,
                        motivo=(
                            f"El total de {etiqueta} del archivo ({del_libro:,}) no coincide "
                            f"con lo cargado ({cargado:,}). Falta alguien por cruzar."
                        ),
                    )
                )
        return avisos

    # ── Interno ───────────────────────────────────────────────────────────────

    @staticmethod
    def _exigir_periodo_abierto(periodo: Periodo) -> None:
        """§7: un período cerrado no admite cambios de presupuesto."""
        if periodo.cerrado:
            raise ErrorPeriodoCerrado(
                f"El período {periodo.codigo} está cerrado y no admite cambios de "
                "presupuesto agropecuario."
            )

    def _historiar(
        self,
        fila: AgroPresupuesto,
        campo: str,
        anterior: Decimal | None,
        nuevo: Decimal | None,
        motivo: str,
        usuario: Usuario | None,
    ) -> None:
        if anterior is not None and anterior == nuevo:
            return
        self._sesion.add(
            AgroPresupuestoHistorial(
                presupuesto_id=fila.id,
                periodo_id=fila.periodo_id,
                dimension=fila.dimension,
                clave=fila.clave,
                campo=campo,
                valor_anterior=anterior,
                valor_nuevo=nuevo,
                motivo=motivo,
                usuario_id=usuario.id if usuario else None,
                cuando=ahora_utc(),
            )
        )

    def _autores(self) -> dict[int | None, str]:
        filas = self._sesion.execute(select(Usuario.id, Usuario.usuario)).all()
        return {fila[0]: fila[1] for fila in filas}

    def _etiquetas_catalogo(self) -> dict[tuple[str, str], str]:
        """`{(tipo, clave): nombre}` de todo el catálogo de dimensiones."""
        filas = self._sesion.execute(
            select(AgroDimension.tipo, AgroDimension.clave, AgroDimension.nombre)
        ).all()
        return {(fila[0], fila[1]): fila[2] for fila in filas}

    def _nombre_catalogo(self, dimension: DimensionPresupuesto, clave: str) -> str:
        return self._etiquetas_catalogo().get((dimension.tipo.value, clave), clave)

    @staticmethod
    def _nombre(
        dimension: DimensionPresupuesto,
        meta: MetaMiembro,
        etiquetas: Mapping[tuple[str, str], str],
    ) -> str:
        """Nombre legible del miembro: catálogo, etiqueta capturada o la clave.

        En ese orden y por este motivo: el catálogo manda porque es lo que dice
        el ERP hoy; la etiqueta con la que se capturó es el respaldo para el
        miembro que todavía no ha facturado y por tanto no está en el catálogo;
        y la clave es el último recurso, que al menos es cierto. Lo que no se
        hace en ningún caso es dejar la fila fuera: una meta sin nombre sigue
        siendo una meta.
        """
        return etiquetas.get((dimension.tipo.value, meta.clave)) or meta.etiqueta or meta.clave


@dataclass(frozen=True, slots=True)
class _FilaCarga:
    """Una fila de la carga masiva, ya validada."""

    dimension: DimensionPresupuesto
    clave: str
    etiqueta: str | None
    monto: Decimal
    kilos: Decimal


def _filas_csv(contenido: bytes) -> list[tuple[int, dict[str, object]]]:
    texto = contenido.decode("utf-8-sig", errors="replace")
    lector = csv.DictReader(io.StringIO(texto))
    # `numero` cuenta desde 2 porque la 1 es el encabezado: así el número que se
    # le devuelve al usuario coincide con el que ve en su editor.
    return [
        (numero, {k: v for k, v in fila.items() if k is not None})
        for numero, fila in enumerate(lector, start=2)
    ]


def _filas_excel(contenido: bytes) -> list[tuple[int, dict[str, object]]]:
    from openpyxl import load_workbook

    libro = load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    try:
        hoja = libro[libro.sheetnames[0]]
        filas_crudas: Iterable[tuple[Any, ...]] = hoja.iter_rows(values_only=True)
        encabezado: list[str] | None = None
        salida: list[tuple[int, dict[str, object]]] = []
        for numero, valores in enumerate(filas_crudas, start=1):
            if encabezado is None:
                encabezado = [str(v) if v is not None else "" for v in valores]
                continue
            salida.append((numero, dict(zip(encabezado, valores, strict=False))))
        return salida
    finally:
        libro.close()


def _normalizar_fila(cruda: Mapping[str, object]) -> _FilaCarga:
    """Una fila del archivo a `_FilaCarga`, o `ErrorValidacion` con su motivo."""
    valores: dict[str, object] = {}
    for titulo, valor in cruda.items():
        destino = _ALIAS_COLUMNAS.get(clave_columna(titulo))
        if destino is not None and valores.get(destino) in (None, ""):
            valores[destino] = valor

    dimension = _dimension_de(valores.get("dimension"))
    clave = normalizar_texto(valores.get("clave"), limite=60)
    if clave is None:
        raise ErrorValidacion(
            "La fila no trae la clave del miembro. Es el identificador del origen "
            "(`CodigoVendedor`, `CO_Id`, `Especie_Id` o `TipoComercial_Id`), y sin él la "
            "meta no se puede cruzar con la venta."
        )

    return _FilaCarga(
        dimension=dimension,
        clave=clave,
        etiqueta=normalizar_etiqueta(valores.get("etiqueta")),
        monto=_a_decimal(valores.get("monto"), "monto", ESCALA_DINERO),
        kilos=_a_decimal(valores.get("kilos"), "kilos", ESCALA_KILOS),
    )


def _dimension_de(valor: object) -> DimensionPresupuesto:
    """Traduce la columna `dimension` del archivo. **No adivina.**

    Una dimensión mal leída pondría la meta de un vendedor en el renglón de una
    especie y descuadraría las dos descomposiciones sin que nadie supiera por
    qué. Por eso lo que no está en la lista de grafías se rechaza nombrando las
    cuatro admitidas, en lugar de asimilarse a la que más se le parezca.
    """
    crudo = normalizar_texto(valor)
    if crudo is None:
        raise ErrorValidacion(
            "La fila no dice a qué dimensión pertenece la meta. El presupuesto agropecuario "
            "se fija por vendedor, centro de operación, especie o tipo comercial, y son "
            "**cuatro repartos del mismo total**: sin la columna «dimension» no se sabe en "
            "cuál de los cuatro va esta fila, y meterla en el que no es descuadraría los dos."
        )
    dimension = _ALIAS_DIMENSION.get(clave_columna(crudo))
    if dimension is None:
        admitidas = ", ".join(d.value for d in DimensionPresupuesto)
        raise ErrorValidacion(
            f"«{crudo}» no es una dimensión de presupuesto agropecuario. Las admitidas son: "
            f"{admitidas}. No se adivina la más parecida: una meta puesta en la dimensión "
            "equivocada descuadra las dos descomposiciones."
        )
    return dimension


def _a_decimal(valor: object, campo: str, escala: Decimal) -> Decimal:
    """Importe del archivo a `Decimal`, sin pasar nunca por `float`."""
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        return CERO.quantize(escala)
    texto = str(valor).strip().replace("$", "").replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        numero = Decimal(texto)
    except (InvalidOperation, ValueError) as exc:
        raise ErrorValidacion(f"El campo «{campo}» no es un número: {valor!r}.") from exc
    if numero < CERO:
        raise ErrorValidacion(f"El campo «{campo}» no puede ser negativo: {valor!r}.")
    return numero.quantize(escala)


_MENSAJE_CUADRA = (
    "Las descomposiciones capturadas dan el mismo total. Es lo que tiene que pasar: son "
    "el mismo dinero repartido de varias formas."
)

_MENSAJE_CUADRE_TRIVIAL = (
    "Solo hay una dimensión con presupuesto capturado, así que no hay con qué contrastar. "
    "No es un descuadre: es un presupuesto a medio parametrizar."
)

_MENSAJE_SIN_PRESUPUESTO = (
    "El período no tiene presupuesto agropecuario capturado en ninguna dimensión. El "
    "cumplimiento viaja vacío y el semáforo sale SIN_PRESUPUESTO, nunca en rojo."
)


def _mensaje_descuadre(
    totales: Mapping[DimensionPresupuesto, tuple[Decimal, Decimal]],
    diferencia_monto: Decimal,
    diferencia_kilos: Decimal,
) -> str:
    """El aviso de descuadre: qué dimensiones, cuánto y qué significa.

    Nombra los totales uno a uno en lugar de decir «no cuadra», porque quien lo
    lea tiene que poder ver de un vistazo cuál de los cuatro repartos es el que
    se desvía y por cuánto.
    """
    detalle = " · ".join(
        f"{dimension.etiqueta}: {monto} en pesos y {kilos} en kilos"
        for dimension, (monto, kilos) in sorted(totales.items(), key=lambda par: par[0].value)
    )
    return (
        "Las descomposiciones del presupuesto NO dan el mismo total, y deberían: son el "
        f"mismo dinero repartido de varias formas. {detalle}. Diferencia máxima: "
        f"{diferencia_monto} en pesos y {diferencia_kilos} en kilos. Es un error de captura "
        "y hay que corregirlo en el origen; el sistema no lo reparte por su cuenta porque "
        "eso sería inventarse la meta de alguien. Mientras tanto el cumplimiento de cada "
        "dimensión sigue calculándose contra su propio presupuesto, que es lo único "
        "correcto que se puede hacer."
    )
