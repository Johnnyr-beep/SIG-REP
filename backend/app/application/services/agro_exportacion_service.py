"""Exportación de los reportes de agropecuaria a Excel.

Mismo criterio que `exportacion_service` y por la misma razón: **se exporta lo
mismo que se muestra**, a partir de la respuesta ya calculada. Recalcular por
otro camino es cómo se llega a que la pantalla y el archivo digan cosas
distintas, que es el problema del que viene huyendo este proyecto.

Los números salen como números —no como texto— para que quien reciba el archivo
pueda seguir trabajando con él, y la fecha de corte va en la primera fila
siempre (§6).

Lo propio de agropecuaria son dos cosas, y las dos consisten en escribir en el
libro algo que la respuesta ya dice y que se perdería al exportar:

**1. La conciliación del impuesto va en el libro.** El `TipoItem = IMPUESTO`
está fuera de todos los totales, así que un archivo sin esa nota no cuadra
contra el ERP y nadie sabe por qué faltan cuarenta millones. Va al pie, con su
importe.

**2. Un cruce truncado lo dice en la hoja, no solo en la pantalla.** El archivo
se reenvía por correo y sobrevive a la sesión que lo generó; si las 500 filas
publicadas suman doscientos millones menos que el consolidado, eso tiene que
viajar dentro del archivo.
"""

from __future__ import annotations

import io
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.schemas.agro import (
    RespuestaCruceAgro,
    RespuestaResumenAgro,
    RespuestaVentaDiariaAgro,
)

#: El orden es el de `IndicadoresAgro`, y se mantiene: la hoja y el esquema se
#: leen en paralelo cuando alguien tiene que comprobar una cifra.
_ENCABEZADOS = [
    "Venta",
    "Venta (pesos)",
    "Kilos",
    "Cantidad",
    "Líneas facturadas",
    "Participación",
    "Margen valor",
    "Margen %",
    "Presupuesto",
    "Cumplimiento",
    "Ideal",
    "Brecha",
    "Semáforo",
    "Proyección",
    "Cumpl. proyectado",
    "Venta diaria promedio",
    "Venta diaria requerida",
    "Días hábiles",
    "Días trabajados",
]

_FORMATO_PORCENTAJE = "0.00%"
_FORMATO_MONEDA = "#,##0.00"
#: Columnas de `_ENCABEZADOS` que son porcentaje (base 0).
_INDICES_PORCENTAJE = {5, 7, 9, 10, 11, 14}


def exportar_resumen(datos: RespuestaResumenAgro) -> bytes:
    libro, hoja = _libro(
        f"Resumen · {datos.por.value}", datos.periodo, str(datos.fecha_corte), datos.medida.value
    )

    hoja.append(["Código", "Nombre", *_ENCABEZADOS])
    _formatear_encabezado(hoja)

    _fila(hoja, ["", "CONSOLIDADO"], datos.consolidado, negrita=True)
    for fila in datos.filas:
        _fila(hoja, [fila.clave, fila.nombre], fila)

    _conciliacion(hoja, datos.parametros_calculo)
    return _cerrar(libro, hoja, columnas=2 + len(_ENCABEZADOS))


def exportar_cruce(datos: RespuestaCruceAgro) -> bytes:
    libro, hoja = _libro(
        f"Cruce · {datos.por.value}", datos.periodo, str(datos.fecha_corte), datos.medida.value
    )

    # Una columna por eje: los dos cruces tienen distinto número y la cabecera
    # se arma de `ejes`, igual que hace la pantalla.
    hoja.append([*[eje.replace("_", " ").capitalize() for eje in datos.ejes], *_ENCABEZADOS])
    _formatear_encabezado(hoja)

    relleno = [""] * (len(datos.ejes) - 1)
    _fila(hoja, ["CONSOLIDADO", *relleno], datos.consolidado, negrita=True)
    for fila in datos.filas:
        _fila(hoja, list(fila.nombres), fila)

    if datos.truncado:
        # El aviso va **dentro** del archivo porque el archivo se reenvía y
        # sobrevive a la pantalla que lo generó. Sin esto, quien lo reciba suma
        # la columna, no le cuadra con el consolidado y no tiene forma de saber
        # que está mirando solo las mayores.
        hoja.append([])
        hoja.append(
            [
                f"Se exportaron las {len(datos.filas)} filas de mayor venta "
                f"(límite: {datos.limite}). El consolidado de arriba es el del corte "
                f"completo, no el de las filas exportadas: por eso la suma de la "
                f"columna de venta es menor."
            ]
        )
        hoja.cell(row=hoja.max_row, column=1).font = Font(bold=True)

    _conciliacion(hoja, datos.parametros_calculo)
    return _cerrar(libro, hoja, columnas=len(datos.ejes) + len(_ENCABEZADOS))


def exportar_venta_diaria(datos: RespuestaVentaDiariaAgro) -> bytes:
    libro, hoja = _libro("Venta diaria", datos.periodo, str(datos.fecha_corte), datos.medida.value)

    hoja.append(["Centro", "Nombre", "Ppto. diario", *[str(f) for f in datos.fechas], "Total"])
    _formatear_encabezado(hoja)

    for fila in datos.filas:
        hoja.append(
            [
                fila.centro,
                fila.nombre,
                datos.presupuesto_diario_por_centro.get(fila.centro),
                *fila.valores,
                fila.total,
            ]
        )

    # La fila de totales es la que publica la respuesta, no una suma hecha aquí:
    # así el archivo no puede discrepar de la pantalla.
    hoja.append(
        [
            "",
            f"TOTAL · {len(datos.filas)} centro(s)",
            datos.totales.presupuesto_diario,
            *datos.totales.valores,
            datos.totales.total,
        ]
    )
    for celda in hoja[hoja.max_row]:
        celda.font = Font(bold=True)

    _conciliacion(hoja, datos.parametros_calculo)
    return _cerrar(libro, hoja, columnas=4 + len(datos.fechas))


# ── Interno ───────────────────────────────────────────────────────────────────


def _libro(titulo: str, periodo: str, fecha_corte: str, medida: str) -> tuple[Workbook, Any]:
    libro = Workbook()
    hoja = libro.active if libro.active is not None else libro.create_sheet()
    hoja.title = titulo[:31]
    hoja.append([f"SIGREP · Agropecuaria · {titulo}"])
    hoja.append([f"Período: {periodo}", f"Fecha de corte: {fecha_corte}", f"Medida: {medida}"])
    hoja.append([])
    hoja["A1"].font = Font(bold=True, size=14)
    return libro, hoja


def _formatear_encabezado(hoja: Any) -> None:
    for celda in hoja[hoja.max_row]:
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center", wrap_text=True)


def _fila(hoja: Any, prefijo: list[str], fila: Any, *, negrita: bool = False) -> None:
    valores: list[Any] = [
        fila.venta,
        fila.venta_valor,
        fila.kilos,
        fila.cantidad,
        fila.lineas_facturadas,
        fila.participacion,
        fila.margen_valor,
        fila.margen_porcentaje,
        fila.presupuesto,
        fila.cumplimiento,
        fila.ideal,
        fila.brecha,
        fila.semaforo.value,
        fila.proyeccion,
        fila.cumplimiento_proyectado,
        fila.venta_diaria_promedio,
        fila.venta_diaria_requerida,
        fila.dias_habiles,
        fila.dias_trabajados,
    ]
    hoja.append([*prefijo, *valores])

    numero_fila = hoja.max_row
    for indice, valor in enumerate(valores):
        celda = hoja.cell(row=numero_fila, column=len(prefijo) + indice + 1)
        if indice in _INDICES_PORCENTAJE:
            celda.number_format = _FORMATO_PORCENTAJE
        elif isinstance(valor, Decimal):
            celda.number_format = _FORMATO_MONEDA
        if negrita:
            celda.font = Font(bold=True)

    if negrita:
        for columna in range(1, len(prefijo) + 1):
            hoja.cell(row=numero_fila, column=columna).font = Font(bold=True)


def _conciliacion(hoja: Any, parametros: Any) -> None:
    """Al pie: lo que se dejó fuera de los totales, y el descuadre si lo hay.

    Sin esto, el archivo no se puede conciliar contra el ERP —falta el impuesto
    y no dice cuánto—, y un descuadre entre las cuatro descomposiciones del
    presupuesto viajaría invisible dentro de un libro que se ve completo.
    """
    conciliacion = parametros.conciliacion
    hoja.append([])
    hoja.append(["Conciliación"])
    hoja.cell(row=hoja.max_row, column=1).font = Font(bold=True)
    hoja.append(
        [
            "Impuesto excluido de todos los totales",
            conciliacion.impuesto_valor,
            f"{conciliacion.impuesto_lineas} líneas facturadas",
        ]
    )
    hoja.cell(row=hoja.max_row, column=2).number_format = _FORMATO_MONEDA
    hoja.append([conciliacion.nota])

    cuadre = parametros.cuadre
    if cuadre is not None and not cuadre.cuadra:
        hoja.append([])
        hoja.append([f"AVISO · {cuadre.mensaje}"])
        hoja.cell(row=hoja.max_row, column=1).font = Font(bold=True)


def _cerrar(libro: Workbook, hoja: Any, *, columnas: int) -> bytes:
    hoja.freeze_panes = "A5"
    for indice in range(1, columnas + 1):
        hoja.column_dimensions[get_column_letter(indice)].width = 18

    memoria = io.BytesIO()
    libro.save(memoria)
    libro.close()
    return memoria.getvalue()
