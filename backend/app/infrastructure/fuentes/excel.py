"""`FuenteVentaExcel`: la venta leída del libro que hoy arma el negocio (§5).

Es la implementación que permite tener SIGREP funcionando y conciliado contra el
Excel vigente antes de que llegue la API de SIESA.

**No se puede cargar el libro entero en memoria.** El archivo real pesa 18 MB y
su hoja `VENTA` tiene 131 819 filas para nueve días —unas 440 000 al mes—, así
que se lee con `openpyxl` en modo `read_only`, que va desmontando el XML fila a
fila, y se devuelve un **generador**. Quien consume decide el tamaño del lote;
esta clase nunca materializa la hoja.

Lo que esta clase hace y lo que no:

- **Sí**: localizar las columnas por su encabezado —con o sin tildes, con o sin
  mayúsculas—, convertir a `Decimal` y a `date`, filtrar por rango de fechas y
  por centro, y **registrar en `rechazos` toda fila que no pudo formar**.
- **No**: normalizar el C.O. a tres posiciones, resolver la categoría contra el
  mapeo ni decidir qué es una clase de cliente válida. Eso es de la ingesta, que
  es quien tiene la base delante (§3.4 y el docstring de `domain/puertos.py`).

La hoja de clientes se llama `'CLIENTES '`, **con un espacio al final**, y **no
tiene fila de encabezado**: la primera fila ya es un cliente. Las dos cosas son
del archivo real y las dos están contempladas aquí.
"""

from __future__ import annotations

import io
from collections.abc import Iterator, Sequence
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from app.core.errors import ErrorValidacion
from app.domain.normalizacion import (
    ESCALA_DINERO,
    ESCALA_KILOS,
    ESCALA_PORCENTAJE,
    a_decimal,
    a_fecha,
    clave_columna,
    normalizar_texto,
)
from app.domain.puertos import LineaVenta
from app.infrastructure.fuentes.base import ClienteFuente, RechazoFuente

#: Encabezados de la hoja `VENTA` reducidos a llave (`clave_columna`) y llevados
#: al nombre con el que viajan dentro de SIGREP. Se aceptan varias grafías
#: porque el encabezado lo escribe quien exporta de SIESA.
#:
#: Nótese que `categoria` y `nueva categoria` son llaves **distintas**: la
#: comparación es por igualdad exacta y no por «contiene». `NUEVA CATEGORIA` es
#: la reclasificación que hoy se hace a mano en el Excel y que en SIGREP produce
#: la tabla `mapeo_categorias`; leerla sería importar la decisión de ayer.
_COLUMNAS_VENTA: dict[str, str] = {
    "c o": "centro_operacion",
    "co": "centro_operacion",
    "centro de operacion": "centro_operacion",
    "fecha": "fecha",
    "cliente pos": "cliente_pos",
    "cliente factura": "cliente_factura",
    "razon social cliente pos": "razon_social_pos",
    "razon social cliente factura": "razon_social_factura",
    "costo promedio": "costo_promedio",
    "cantidad inv": "cantidad_inv",
    "valor subtotal": "valor_subtotal",
    "margen": "margen",
    "domicilio": "domicilio",
    "clases de clientes": "clase_cliente",
    "clase de cliente": "clase_cliente",
    "condicion de pago": "condicion_pago",
    "categoria": "categoria_siesa",
}

#: Sin estas tres no hay línea de venta que valga: el punto de venta, el día y
#: la medida contra la que se compara el presupuesto.
_COLUMNAS_OBLIGATORIAS = ("centro_operacion", "fecha", "valor_subtotal")

NOMBRE_HOJA_VENTA = "VENTA"
NOMBRE_HOJA_CLIENTES = "CLIENTES"

_CERO = Decimal("0")


class FuenteVentaExcel:
    """Implementación del puerto `FuenteVenta` sobre el libro de SIESA."""

    def __init__(self, origen: str | Path | bytes, *, nombre: str | None = None) -> None:
        """`origen` es una ruta en disco o el contenido de un `.xlsx` subido."""
        self._origen = origen
        self._nombre = nombre or (
            Path(origen).name if isinstance(origen, str | Path) else "venta.xlsx"
        )
        self._libro: Workbook | None = None
        #: Filas que no se pudieron convertir en `LineaVenta`. La ingesta las
        #: recoge al terminar y las vuelca en la bitácora de la corrida.
        self.rechazos: list[RechazoFuente] = []

    # ── Identidad ─────────────────────────────────────────────────────────────

    @property
    def nombre(self) -> str:
        """Nombre del origen. Va a `corridas_ingesta.origen`."""
        return self._nombre

    @property
    def hojas(self) -> set[str]:
        """Nombres de hoja normalizados (sin espacios sobrantes, en mayúscula)."""
        return {str(nombre).strip().upper() for nombre in self._abrir().sheetnames}

    @property
    def tiene_venta(self) -> bool:
        return NOMBRE_HOJA_VENTA in self.hojas

    @property
    def tiene_clientes(self) -> bool:
        return NOMBRE_HOJA_CLIENTES in self.hojas

    def cerrar(self) -> None:
        """Libera el libro. En `read_only` es obligatorio: deja ficheros abiertos."""
        if self._libro is not None:
            self._libro.close()
            self._libro = None

    # ── Puerto `FuenteVenta` ──────────────────────────────────────────────────

    def obtener_ventas(
        self,
        desde: date,
        hasta: date,
        centros: Sequence[str] | None = None,
    ) -> Iterator[LineaVenta]:
        """Líneas de la hoja `VENTA` del rango, ambos extremos incluidos.

        Es un generador: no hay una lista de 131 000 elementos en ninguna parte.
        Las filas fuera del rango o de los centros pedidos **no se cuentan como
        leídas** —no forman parte de esta corrida—; las que sí caen dentro y no
        se pudieron formar sí, y salen por `rechazos`.

        La fila con la fecha ilegible se rechaza aunque se esté pidiendo un rango
        estrecho: no se puede saber si cae dentro, y una fila que se descarta por
        no saber en qué día va es exactamente lo que §5 prohíbe callarse.
        """
        hoja = self._hoja(NOMBRE_HOJA_VENTA)
        if hoja is None:
            return

        filas = hoja.iter_rows(values_only=True)
        try:
            encabezado = next(filas)
        except StopIteration:
            return

        columnas = self._mapear_columnas(encabezado)
        pedidos = {str(c).strip().zfill(3) for c in centros} if centros else None

        for numero, valores in enumerate(filas, start=2):
            if all(valor is None for valor in valores):
                continue

            fecha = a_fecha(self._celda(valores, columnas, "fecha"))
            if fecha is None:
                self._rechazar(numero, "Fecha", self._celda(valores, columnas, "fecha"))
                continue
            if fecha < desde or fecha > hasta:
                continue

            centro = normalizar_texto(self._celda(valores, columnas, "centro_operacion"))
            if centro is None:
                self._rechazar(numero, "C.O.", None, "La fila no indica centro de operación.")
                continue
            if pedidos is not None and centro.zfill(3) not in pedidos:
                continue

            linea = self._formar_linea(numero, valores, columnas, centro, fecha)
            if linea is not None:
                yield linea

    # ── Catálogo de clientes ──────────────────────────────────────────────────

    def obtener_clientes(self) -> Iterator[ClienteFuente]:
        """Registros de la hoja `CLIENTES`: NIT, razón social, canal y vendedor.

        **La hoja no tiene encabezado**: la primera fila del archivo real ya es
        un cliente (`901733567 · ALAMO EDUCA SAS · HORECA · Michel Barrios`).
        Aun así se comprueba la primera fila: si su primera celda no es un
        número, es que alguien le añadió encabezado y se salta. Perder 451
        clientes por un `NIT` en la fila 1 sería absurdo, y perder uno por
        suponer que siempre habrá encabezado, también.
        """
        hoja = self._hoja(NOMBRE_HOJA_CLIENTES)
        if hoja is None:
            return

        for numero, valores in enumerate(hoja.iter_rows(values_only=True), start=1):
            if not valores or all(valor is None for valor in valores):
                continue
            crudo = normalizar_texto(valores[0])
            if crudo is None:
                continue
            if numero == 1 and not crudo.replace(".0", "").isdigit():
                continue  # la fila 1 resultó ser un encabezado
            yield ClienteFuente(
                nit=valores[0],
                razon_social=valores[1] if len(valores) > 1 else None,
                canal=valores[2] if len(valores) > 2 else None,
                vendedor=valores[3] if len(valores) > 3 else None,
                fila_origen=numero,
            )

    # ── Interno ───────────────────────────────────────────────────────────────

    def _abrir(self) -> Workbook:
        if self._libro is None:
            fuente: Any = (
                io.BytesIO(self._origen) if isinstance(self._origen, bytes) else self._origen
            )
            try:
                self._libro = load_workbook(fuente, read_only=True, data_only=True)
            except Exception as exc:  # openpyxl lanza de todo ante un archivo roto
                raise ErrorValidacion(f"No se pudo abrir el libro {self._nombre!r}: {exc}") from exc
        return self._libro

    def _hoja(self, nombre: str) -> Any:
        """La hoja por su nombre normalizado.

        La hoja de clientes del archivo real se llama `'CLIENTES '`, con un
        espacio al final. Buscar por el nombre exacto habría costado un
        `KeyError` en producción y media tarde de depuración.
        """
        libro = self._abrir()
        for real in libro.sheetnames:
            if str(real).strip().upper() == nombre:
                return libro[real]
        return None

    def _mapear_columnas(self, encabezado: Sequence[object]) -> dict[str, int]:
        columnas: dict[str, int] = {}
        for indice, titulo in enumerate(encabezado):
            campo = _COLUMNAS_VENTA.get(clave_columna(titulo))
            if campo is not None and campo not in columnas:
                columnas[campo] = indice

        faltantes = [c for c in _COLUMNAS_OBLIGATORIAS if c not in columnas]
        if faltantes:
            raise ErrorValidacion(
                f"La hoja {NOMBRE_HOJA_VENTA} de {self._nombre!r} no trae las columnas "
                "obligatorias: " + ", ".join(faltantes) + ". "
                "Se esperan al menos «C.O.», «Fecha» y «Valor subtotal»."
            )
        return columnas

    @staticmethod
    def _celda(valores: Sequence[object], columnas: dict[str, int], campo: str) -> object:
        indice = columnas.get(campo)
        if indice is None or indice >= len(valores):
            return None
        return valores[indice]

    def _formar_linea(
        self,
        numero: int,
        valores: Sequence[object],
        columnas: dict[str, int],
        centro: str,
        fecha: date,
    ) -> LineaVenta | None:
        """Convierte una fila en `LineaVenta`, o la rechaza con su motivo.

        Una celda **vacía** de importe vale cero: no hay nada que interpretar y
        rechazar la fila costaría una venta real. Una celda **con contenido que
        no es un número** sí rechaza la fila: ahí sí hay algo que alguien tiene
        que mirar, y convertirlo en cero sería inventarse el dato.
        """
        medidas: dict[str, Decimal] = {}
        for campo, etiqueta, escala in (
            ("valor_subtotal", "Valor subtotal", ESCALA_DINERO),
            ("costo_promedio", "Costo promedio", ESCALA_DINERO),
            ("cantidad_inv", "Cantidad inv.", ESCALA_KILOS),
        ):
            crudo = self._celda(valores, columnas, campo)
            if normalizar_texto(crudo) is None:
                medidas[campo] = _CERO
                continue
            numero_decimal = a_decimal(crudo, escala)
            if numero_decimal is None:
                self._rechazar(numero, etiqueta, crudo, f"«{etiqueta}» no es un número.")
                return None
            medidas[campo] = numero_decimal

        nit = self._celda(valores, columnas, "cliente_factura")
        razon = self._celda(valores, columnas, "razon_social_factura")
        if normalizar_texto(nit) is None:
            # Sin cliente de factura manda el del POS: es el mismo en el 100 %
            # de las filas del archivo actual, pero el orden importa el día que
            # dejen de serlo —quien factura es quien identifica al comprador—.
            nit = self._celda(valores, columnas, "cliente_pos")
            razon = self._celda(valores, columnas, "razon_social_pos")

        return LineaVenta(
            centro_operacion=centro,
            fecha=fecha,
            valor_subtotal=medidas["valor_subtotal"],
            costo_promedio=medidas["costo_promedio"],
            cantidad_inv=medidas["cantidad_inv"],
            categoria_siesa=normalizar_texto(self._celda(valores, columnas, "categoria_siesa")),
            nit_cliente=normalizar_texto(nit),
            razon_social=normalizar_texto(razon),
            margen_siesa=a_decimal(self._celda(valores, columnas, "margen"), ESCALA_PORCENTAJE),
            domicilio=normalizar_texto(self._celda(valores, columnas, "domicilio")),
            clase_cliente=normalizar_texto(self._celda(valores, columnas, "clase_cliente")),
            condicion_pago=normalizar_texto(self._celda(valores, columnas, "condicion_pago")),
            fila_origen=numero,
        )

    def _rechazar(self, fila: int, campo: str, valor: object, motivo: str | None = None) -> None:
        self.rechazos.append(
            RechazoFuente(
                fila=fila,
                campo=campo,
                valor=normalizar_texto(valor, limite=300),
                motivo=motivo or f"«{campo}» no se pudo interpretar.",
            )
        )
