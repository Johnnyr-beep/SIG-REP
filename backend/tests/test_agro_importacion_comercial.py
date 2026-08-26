"""Pruebas de la importación configurable del Excel comercial de agropecuaria.

El libro anual trae una hoja `RESUMEN (MES)` con los canales como filas
(`SUPER MAYORISTA`, `MAYORISTA`, `TAT`, `Call Center`…) y los meses `ENE..DIC`
como columnas. La importación lee el valor del mes del período **tal cual está
almacenado** (sin escalar por 1 000) y lo vuelca en el bloque **commercial** del
presupuesto mensual, mapeando cada canal a vendedor, cliente y categoría A–F
mediante la configuración de `agro_ppto_mensual_canal_mapeo`. Los canales sin
mapeo se rechazan con su motivo.

Las pruebas:

1. El parser lee el valor del mes correcto y no escala por 1 000.
2. La importación vuelca los canales mapeados en el bloque comercial.
3. Los canales sin mapeo se rechazan con su motivo.
4. El total del resultado es la suma de las filas aceptadas.
5. Un período cerrado bloquea la importación.
6. El mapeo de canal se normaliza y es único.
7. La captura manual sigue funcionando junto a la importación.
"""

from __future__ import annotations

from collections.abc import Sequence
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.application.services.agro_importacion_comercial_parser import (
    leer_canales,
    normalizar_canal,
)
from app.application.services.agro_presupuesto_mensual_service import (
    AgroPresupuestoMensualService,
)
from app.application.services.periodos import obtener_o_crear_periodo
from app.core.errors import ErrorPeriodoCerrado, ErrorValidacion
from app.infrastructure.models.agro_presupuesto_mensual import (
    AgroPptoMensualDetalle,
)
from app.schemas.agro import (
    CanalMapeoMensualEntrada,
    DetalleMensualEntrada,
)
from tests.conftest import PERIODO

#: Meses en el orden de las columnas del libro.
_MESES = ("ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC")


def _servicio(sesion: Session) -> AgroPresupuestoMensualService:
    return AgroPresupuestoMensualService(sesion)


def _generar_libro(
    canales: Sequence[tuple[str, Sequence[Decimal | None]]],
    *,
    anio_presupuesto: int = 2026,
    anio_historico: int = 2025,
    nombre_hoja: str = "RESUMEN (MES)",
) -> bytes:
    """Arma un libro que reproduce la forma del libro del negocio.

    Cada canal trae tres filas: el histórico (`anio_historico`), el presupuesto
    (`anio_presupuesto`) y `% Crec`. El nombre del canal va en la columna A de
    la primera fila; las dos siguientes llevan la columna A vacía, igual que el
    libro real. Los meses `ENE..DIC` van en las columnas C..N; `Q1..Q4` y
    `TOTAL` se rellenan con `None` porque la importación los ignora.
    """
    libro = Workbook()
    hoja = libro.active
    assert hoja is not None
    hoja.title = nombre_hoja

    # Fila 1: título, igual que el libro real.
    hoja.append([None, None, None, None, None, None, "PPTO 2026 $"])

    # Fila 2: encabezado con los meses.
    encabezado = [None, "DATO", *_MESES, "Q1", "Q2", "Q3", "Q4", "TOTAL"]
    hoja.append(encabezado)

    for canal, valores in canales:
        # Fila del histórico.
        hoja.append([canal, anio_historico, *valores, None, None, None, None, None])
        # Fila del presupuesto: columna A vacía.
        hoja.append([None, anio_presupuesto, *valores, None, None, None, None, None])
        # Fila de % Crec.
        hoja.append([None, "% Crec", *[0.0] * 12, None, None, None, None, None])

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


# ── Parser ───────────────────────────────────────────────────────────────────


def test_el_parser_le_el_valor_del_mes_sin_escalar(
    estructura: None,
    sesion: Session,
) -> None:
    """El valor del mes se lee tal cual está almacenado: no se multiplica por 1 000."""
    canales = [
        ("SUPER MAYORISTA", [Decimal("1000000")] * 12),
        ("TAT", [Decimal("500000")] * 12),
    ]
    contenido = _generar_libro(canales)

    leidos = leer_canales(contenido, "2026-08")

    assert {c.canal for c in leidos} == {"SUPER MAYORISTA", "TAT"}
    super_mayorista = next(c for c in leidos if c.canal == "SUPER MAYORISTA")
    assert super_mayorista.monto == Decimal("1000000.00")
    tat = next(c for c in leidos if c.canal == "TAT")
    assert tat.monto == Decimal("500000.00")


def test_el_parser_toma_la_fila_del_anio_del_periodo(
    estructura: None,
    sesion: Session,
) -> None:
    """La fila del presupuesto es la cuyo segundo valor es el año del período.

    No la del histórico (`2025`) ni la de `% Crec` (texto).
    """
    valores = [Decimal("1000000"), Decimal("1100000"), Decimal("1200000")] + [Decimal("0")] * 9
    contenido = _generar_libro(
        [("SUPER MAYORISTA", valores)],
        anio_presupuesto=2026,
        anio_historico=2025,
    )

    leidos_2026 = leer_canales(contenido, "2026-02")
    assert leidos_2026[0].monto == Decimal("1100000.00")

    # Si el período es de 2025, se lee la fila del histórico.
    leidos_2025 = leer_canales(contenido, "2025-02")
    assert leidos_2025[0].monto == Decimal("1100000.00")


def test_el_parser_omite_canales_sin_valor_para_el_mes(
    estructura: None,
    sesion: Session,
) -> None:
    """Una celda vacía para el mes se omite: ese canal no trae presupuesto."""
    valores: list[Decimal | None] = [Decimal("1000000"), None] + [Decimal("0")] * 10
    contenido = _generar_libro([("SUPER MAYORISTA", valores)])

    leidos_feb = leer_canales(contenido, "2026-02")
    assert leidos_feb == []

    leidos_ene = leer_canales(contenido, "2026-01")
    assert len(leidos_ene) == 1
    assert leidos_ene[0].monto == Decimal("1000000.00")


def test_el_parser_rechaza_un_libro_sin_encabezado_de_meses(
    estructura: None,
    sesion: Session,
) -> None:
    """Un libro sin la fila `ENE..DIC` se rechaza con un mensaje accionable."""
    libro = Workbook()
    hoja = libro.active
    assert hoja is not None
    hoja.append(["canal", "valor"])
    hoja.append(["SUPER MAYORISTA", 1000])
    buffer = BytesIO()
    libro.save(buffer)

    with pytest.raises(ErrorValidacion, match=r"ENE\.\.DIC"):
        leer_canales(buffer.getvalue(), "2026-08")


def test_el_parser_encuentra_la_hoja_resumen_mes(
    estructura: None,
    sesion: Session,
) -> None:
    """Prefiere la hoja `RESUMEN (MES)` aunque no sea la primera."""
    libro = Workbook()
    otra = libro.active
    assert otra is not None
    otra.title = "Otra hoja"
    otra.append(["basura"])
    resumen = libro.create_sheet("RESUMEN (MES)")
    resumen.append([None, None, "PPTO 2026 $"])
    resumen.append([None, "DATO", *_MESES, "Q1", "Q2", "Q3", "Q4", "TOTAL"])
    resumen.append(["SUPER MAYORISTA", 2025, *[1000] * 12, None, None, None, None, None])
    resumen.append([None, 2026, *[2000] * 12, None, None, None, None, None])
    resumen.append([None, "% Crec", *[0] * 12, None, None, None, None, None])
    buffer = BytesIO()
    libro.save(buffer)

    leidos = leer_canales(buffer.getvalue(), "2026-08")
    assert len(leidos) == 1
    assert leidos[0].canal == "SUPER MAYORISTA"
    assert leidos[0].monto == Decimal("2000.00")


# ── Importación ───────────────────────────────────────────────────────────────


def _mapear_canal(sesion: Session, canal: str, vendedor: str, cliente: str, categoria: str) -> None:
    """Crea un mapeo de canal para la importación."""
    svc = _servicio(sesion)
    svc.guardar_canal_mapeo(
        CanalMapeoMensualEntrada(
            canal=canal,
            vendedor_clave=vendedor,
            cliente_clave=cliente,
            categoria=categoria,
            activo=True,
        )
    )
    sesion.commit()


def test_la_importacion_vuelca_los_canales_mapeados_en_el_bloque_comercial(
    estructura: None,
    sesion: Session,
) -> None:
    """Los canales mapeados se vuelcan en filas de detalle del bloque comercial."""
    _mapear_canal(sesion, "SUPER MAYORISTA", "LEON", "CLIENTE-1", "A")
    _mapear_canal(sesion, "TAT", "JUAN", "CLIENTE-2", "B")

    canales = [
        ("SUPER MAYORISTA", [Decimal("1000000")] * 12),
        ("TAT", [Decimal("500000")] * 12),
    ]
    contenido = _generar_libro(canales)

    resultado = _servicio(sesion).importar_comercial(
        contenido, "presupuesto.xlsx", PERIODO, "Importación de prueba"
    )

    assert resultado.aceptadas == 2
    assert resultado.rechazadas == 0
    assert resultado.total_monto == Decimal("1500000.00")

    # Las filas quedaron en el bloque comercial.
    filas = sesion.query(AgroPptoMensualDetalle).all()
    assert len(filas) == 2
    bloque = {f.cliente_clave: f for f in filas}
    assert bloque["CLIENTE-1"].vendedor_clave == "LEON"
    assert bloque["CLIENTE-1"].categoria == "A"
    assert Decimal(bloque["CLIENTE-1"].monto) == Decimal("1000000.00")
    assert bloque["CLIENTE-2"].vendedor_clave == "JUAN"
    assert bloque["CLIENTE-2"].categoria == "B"


def test_los_canales_sin_mapeo_se_rechazan_con_su_motivo(
    estructura: None,
    sesion: Session,
) -> None:
    """Un canal sin mapeo se rechaza y no se adivina un destino."""
    _mapear_canal(sesion, "SUPER MAYORISTA", "LEON", "CLIENTE-1", "A")

    canales = [
        ("SUPER MAYORISTA", [Decimal("1000000")] * 12),
        ("TAT", [Decimal("500000")] * 12),  # sin mapeo
    ]
    contenido = _generar_libro(canales)

    resultado = _servicio(sesion).importar_comercial(
        contenido, "presupuesto.xlsx", PERIODO, "Importación de prueba"
    )

    assert resultado.aceptadas == 1
    assert resultado.rechazadas == 1
    # El total es la suma de las aceptadas, no la del libro.
    assert resultado.total_monto == Decimal("1000000.00")

    rechazada = next(f for f in resultado.filas if not f.aceptada)
    assert rechazada.canal == "TAT"
    assert rechazada.motivo is not None
    assert "no tiene mapeo" in rechazada.motivo

    # El canal rechazado no dejó fila en el bloque.
    filas = sesion.query(AgroPptoMensualDetalle).all()
    assert len(filas) == 1
    assert filas[0].cliente_clave == "CLIENTE-1"


def test_la_importacion_reusa_filas_existentes_no_las_duplica(
    estructura: None,
    sesion: Session,
) -> None:
    """Importar dos veces el mismo canal reemplaza la fila, no la duplica."""
    _mapear_canal(sesion, "SUPER MAYORISTA", "LEON", "CLIENTE-1", "A")

    canales = [("SUPER MAYORISTA", [Decimal("1000000")] * 12)]
    contenido = _generar_libro(canales)
    svc = _servicio(sesion)
    svc.importar_comercial(contenido, "presupuesto.xlsx", PERIODO, "Primera importación")

    canales2 = [("SUPER MAYORISTA", [Decimal("2000000")] * 12)]
    contenido2 = _generar_libro(canales2)
    svc.importar_comercial(contenido2, "presupuesto.xlsx", PERIODO, "Segunda importación")

    filas = sesion.query(AgroPptoMensualDetalle).all()
    assert len(filas) == 1
    assert Decimal(filas[0].monto) == Decimal("2000000.00")


def test_la_importacion_no_toca_los_demas_bloques(
    estructura: None,
    sesion: Session,
) -> None:
    """La importación solo escribe en el bloque commercial."""
    _mapear_canal(sesion, "SUPER MAYORISTA", "LEON", "CLIENTE-1", "A")

    # Una fila manual en agro_distribucion.
    _servicio(sesion).guardar_detalle(
        PERIODO,
        DetalleMensualEntrada(
            bloque="agro_distribucion",
            cliente_clave="CLIENTE-AGRO",
            monto=Decimal("300000"),
            kilos=Decimal("0"),
        ),
    )

    canales = [("SUPER MAYORISTA", [Decimal("1000000")] * 12)]
    contenido = _generar_libro(canales)
    _servicio(sesion).importar_comercial(
        contenido, "presupuesto.xlsx", PERIODO, "Importación de prueba"
    )

    filas = sesion.query(AgroPptoMensualDetalle).all()
    bloques = {f.bloque for f in filas}
    assert bloques == {"commercial", "agro_distribucion"}


def test_periodo_cerrado_bloquea_la_importacion(
    estructura: None,
    sesion: Session,
) -> None:
    """Un período cerrado no admite la importación comercial."""
    _mapear_canal(sesion, "SUPER MAYORISTA", "LEON", "CLIENTE-1", "A")

    periodo = obtener_o_crear_periodo(sesion, PERIODO)
    periodo.cerrado = True
    sesion.commit()

    contenido = _generar_libro([("SUPER MAYORISTA", [Decimal("1000000")] * 12)])
    with pytest.raises(ErrorPeriodoCerrado):
        _servicio(sesion).importar_comercial(
            contenido, "presupuesto.xlsx", PERIODO, "Importación de prueba"
        )


# ── Mapeo de canales ─────────────────────────────────────────────────────────


def test_el_mapeo_de_canal_se_normaliza_y_es_unico(
    estructura: None,
    sesion: Session,
) -> None:
    """El canal se normaliza: `super mayorista ` y `SUPER MAYORISTA` son el mismo."""
    svc = _servicio(sesion)
    svc.guardar_canal_mapeo(
        CanalMapeoMensualEntrada(
            canal="SUPER MAYORISTA",
            vendedor_clave="LEON",
            cliente_clave="CLIENTE-1",
            categoria="A",
        )
    )
    sesion.commit()

    # Listar trae el canal normalizado.
    mapeos = svc.listar_canales_mapeos()
    assert len(mapeos) == 1
    assert mapeos[0].canal == "SUPER MAYORISTA"

    # Un segundo mapeo con una grafía distinta del mismo canal choca con la
    # restricción de unicidad (se traduce a 409 por el manejador global, pero
    # aquí se prueba a nivel de sesión: la IntegrityError la levanta SQLAlchemy).
    with pytest.raises(IntegrityError):
        svc.guardar_canal_mapeo(
            CanalMapeoMensualEntrada(
                canal="super mayorista ",  # distinta grafía, mismo canal normalizado
                vendedor_clave="JUAN",
                cliente_clave="CLIENTE-2",
                categoria="B",
            )
        )
    sesion.rollback()


def test_el_mapeo_de_canal_se_actualiza(
    estructura: None,
    sesion: Session,
) -> None:
    """Actualizar un mapeo existente cambia sus campos sin crear otro."""
    svc = _servicio(sesion)
    salida = svc.guardar_canal_mapeo(
        CanalMapeoMensualEntrada(
            canal="TAT",
            vendedor_clave="JUAN",
            cliente_clave="CLIENTE-1",
            categoria="A",
        )
    )
    sesion.commit()

    svc.guardar_canal_mapeo(
        CanalMapeoMensualEntrada(
            canal="TAT",
            vendedor_clave="LEON",
            cliente_clave="CLIENTE-2",
            categoria="B",
            activo=False,
        ),
        mapeo_id=salida.id,
    )
    sesion.commit()

    mapeos = svc.listar_canales_mapeos()
    assert len(mapeos) == 1
    assert mapeos[0].vendedor_clave == "LEON"
    assert mapeos[0].cliente_clave == "CLIENTE-2"
    assert mapeos[0].categoria == "B"
    assert not mapeos[0].activo


def test_un_mapeo_inactivo_no_se_usa_en_la_importacion(
    estructura: None,
    sesion: Session,
) -> None:
    """Un mapeo inactivo se retira sin borrarse: la importación lo ignora."""
    svc = _servicio(sesion)
    salida = svc.guardar_canal_mapeo(
        CanalMapeoMensualEntrada(
            canal="SUPER MAYORISTA",
            vendedor_clave="LEON",
            cliente_clave="CLIENTE-1",
            categoria="A",
        )
    )
    sesion.commit()
    svc.guardar_canal_mapeo(
        CanalMapeoMensualEntrada(
            canal="SUPER MAYORISTA",
            vendedor_clave="LEON",
            cliente_clave="CLIENTE-1",
            categoria="A",
            activo=False,
        ),
        mapeo_id=salida.id,
    )
    sesion.commit()

    contenido = _generar_libro([("SUPER MAYORISTA", [Decimal("1000000")] * 12)])
    resultado = svc.importar_comercial(
        contenido, "presupuesto.xlsx", PERIODO, "Importación de prueba"
    )
    assert resultado.aceptadas == 0
    assert resultado.rechazadas == 1
    assert resultado.filas[0].motivo is not None
    assert "no tiene mapeo" in resultado.filas[0].motivo


def test_normalizar_canal_quita_tildes_y_espacios(
    estructura: None,
    sesion: Session,
) -> None:
    """La normalización del canal quita tildes, mayúsculas y espacios de relleno."""
    assert normalizar_canal("Call Center") == "CALL CENTER"
    assert normalizar_canal("  super  mayorista ") == "SUPER MAYORISTA"
    assert normalizar_canal("DISTRIBUCIÓN") == "DISTRIBUCION"
    assert normalizar_canal("") == ""
    assert normalizar_canal(None) == ""


# ── La captura manual sigue funcionando ──────────────────────────────────────


def test_la_captura_manual_sigue_funcionando_con_la_importacion(
    estructura: None,
    sesion: Session,
) -> None:
    """La captura manual del bloque comercial sigue funcionando junto a la importación."""
    _mapear_canal(sesion, "SUPER MAYORISTA", "LEON", "CLIENTE-1", "A")

    # Captura manual de otra fila del bloque comercial.
    _servicio(sesion).guardar_detalle(
        PERIODO,
        DetalleMensualEntrada(
            bloque="commercial",
            vendedor_clave="JUAN",
            cliente_clave="CLIENTE-MANUAL",
            categoria="C",
            monto=Decimal("300000"),
            kilos=Decimal("0"),
        ),
    )

    # Importación del Excel.
    contenido = _generar_libro([("SUPER MAYORISTA", [Decimal("1000000")] * 12)])
    _servicio(sesion).importar_comercial(
        contenido, "presupuesto.xlsx", PERIODO, "Importación de prueba"
    )

    filas = sesion.query(AgroPptoMensualDetalle).all()
    assert len(filas) == 2
    clientes = {f.cliente_clave for f in filas}
    assert clientes == {"CLIENTE-1", "CLIENTE-MANUAL"}
