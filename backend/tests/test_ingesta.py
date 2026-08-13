"""Ingesta de venta: normalización, idempotencia y bitácora (§3.4, §5, §7).

Cada prueba de este archivo corresponde a un problema **medido** en el archivo
real de SIESA (131 819 filas, 17 columnas, del 2026-08-01 al 2026-08-09), no a
un caso imaginado. El libro de 18 MB no está versionado y no puede estarlo, así
que aquí se construye con `openpyxl` un `.xlsx` mínimo que reproduce cada vicio:
el C.O. que unas veces es texto y otras número, los NIT con relleno, la columna
de clases de cliente corrupta, el domicilio en blanco.

La prueba que más vale es `test_recargar_el_mismo_rango_reemplaza_y_no_duplica`
y su hermana `test_si_la_carga_falla_a_mitad_no_queda_medio_dia_cargado`: la
idempotencia es la regla crítica de §5 y la única cuyo incumplimiento no se nota
—el reporte simplemente sale con el doble de venta y todo el mundo se lo cree—.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.application.services.ingesta_service import IngestaService
from app.domain.enums import EstadoCorrida
from app.domain.normalizacion import (
    a_decimal,
    a_fecha,
    normalizar_centro_operacion,
    normalizar_clase_cliente,
    normalizar_domicilio,
    normalizar_nit,
)
from app.infrastructure.models.catalogo import Categoria
from app.infrastructure.models.ingesta import CorridaIngesta
from app.infrastructure.models.organizacion import PuntoVenta
from app.infrastructure.models.venta import Cliente, VentaLinea
from app.schemas.ingesta import CorridaSalida
from tests.conftest import PERIODO, id_punto_venta

D = Decimal

#: Encabezado exacto de la hoja `VENTA`, con la columna `Columna1` incluida:
#: está en el archivo real entre `CATEGORIA` y `NUEVA CATEGORIA` y la ingesta
#: tiene que ignorarla sin despeinarse.
ENCABEZADO_VENTA = [
    "C.O.",
    "Desc. C.O.",
    "Fecha",
    "Cliente POS",
    "Razón social cliente POS",
    "Cliente factura",
    "Razón social cliente factura",
    "Costo promedio",
    "Cantidad inv.",
    "Valor subtotal",
    "MARGEN",
    "Domicilio",
    "CLASES DE CLIENTES",
    "Condición de pago",
    "CATEGORIA",
    "Columna1",
    "NUEVA CATEGORIA",
]


def fila_venta(
    co: object = "402",
    fecha: object = datetime(2026, 8, 3),
    *,
    nit: object = "900983334      ",
    costo: object = 100.0,
    kilos: object = 2.5,
    valor: object = 500.0,
    margen: object = 20.0,
    domicilio: object = "Si",
    clase: object = "002 - CLIENTES NACIONALES",
    condicion: object = "CON",
    categoria: object = "0001 - RES",
) -> list[object]:
    """Una fila de la hoja `VENTA` con los valores por defecto sanos."""
    return [
        co,
        "PDV MALAMBO",
        fecha,
        nit,
        "CLIENTE DE PRUEBA",
        nit,
        "CLIENTE DE PRUEBA",
        costo,
        kilos,
        valor,
        margen,
        domicilio,
        clase,
        condicion,
        categoria,
        "RES",
        "RES",
    ]


def libro_venta(
    filas: list[list[object]],
    clientes: list[list[object]] | None = None,
    *,
    con_hoja_venta: bool = True,
) -> bytes:
    """Un `.xlsx` con la hoja `VENTA` y, si se pide, la de clientes.

    La hoja de clientes se llama `'CLIENTES '` **con espacio al final** y **no
    lleva encabezado**, igual que en el archivo real. Las dos cosas han costado
    tiempo a alguien alguna vez.
    """
    libro = Workbook()
    hoja: Any = libro.active
    if con_hoja_venta:
        hoja.title = "VENTA"
        hoja.append(ENCABEZADO_VENTA)
        for fila in filas:
            hoja.append(fila)
    else:
        hoja.title = "OTRA COSA"

    if clientes is not None:
        hoja_clientes: Any = libro.create_sheet("CLIENTES ")
        for cliente in clientes:
            hoja_clientes.append(cliente)

    memoria = io.BytesIO()
    libro.save(memoria)
    return memoria.getvalue()


def ingerir(sesion: Session, filas: list[list[object]], **kwargs: Any) -> CorridaSalida:
    salida = IngestaService(sesion).ingerir_archivo(libro_venta(filas, **kwargs), "venta.xlsx")
    sesion.commit()
    return salida


def lineas(sesion: Session) -> list[VentaLinea]:
    return list(sesion.scalars(select(VentaLinea).order_by(VentaLinea.id)))


def motivos(sesion: Session, corrida_id: int, campo: str) -> list[str]:
    return [r.motivo for r in IngestaService(sesion).rechazos(corrida_id) if r.campo == campo]


# ── §3.4 Normalización, función a función ─────────────────────────────────────


@pytest.mark.parametrize(
    ("crudo", "esperado"),
    [
        ("606", "606"),
        (606, "606"),  # así conviven en el archivo real: 1 466 filas y 109
        (606.0, "606"),  # y así llegaría si alguien lo pasa por una hoja de cálculo
        ("  606  ", "606"),
        ("42", "042"),  # tres posiciones con ceros a la izquierda
        (None, None),
        ("   ", None),
    ],
)
def test_el_co_se_normaliza_a_tres_posiciones_venga_como_venga(
    crudo: object, esperado: str | None
) -> None:
    assert normalizar_centro_operacion(crudo) == esperado


def test_el_nit_pierde_el_relleno_de_siesa() -> None:
    """`'900983334      '` en la venta y `900983334` en el catálogo son el mismo."""
    assert normalizar_nit("900983334      ") == "900983334"
    assert normalizar_nit(900983334) == "900983334"
    assert normalizar_nit(" ") is None


@pytest.mark.parametrize(
    "valido",
    ["001 - EMPLEADOS", "002 - CLIENTES NACIONALES", "CONSUMIDOR FINAL PDV ALAMEDA 1"],
)
def test_las_clases_de_cliente_del_catalogo_se_aceptan(valido: str) -> None:
    assert normalizar_clase_cliente(valido) == valido


@pytest.mark.parametrize(
    "basura",
    ["johana.muñoz", "2026-08-03 16:29:02", " ", "", "-", "x"],
)
def test_lo_que_no_es_una_clase_de_cliente_no_pasa(basura: str) -> None:
    assert normalizar_clase_cliente(basura) is None


def test_domicilio_en_blanco_es_nulo_y_no_falso() -> None:
    """Nadie afirmó que esas 28 ventas no fueran a domicilio (§3.4)."""
    assert normalizar_domicilio("Si") is True
    assert normalizar_domicilio("No") is False
    assert normalizar_domicilio(" ") is None
    assert normalizar_domicilio(None) is None


def test_los_importes_se_convierten_sin_pasar_por_float() -> None:
    """`Decimal(0.1)` arrastra el error binario; `Decimal('0.1')` no."""
    assert a_decimal(16366867, D("0.01")) == D("16366867.00")
    assert a_decimal(370.83, D("0.001")) == D("370.830")
    assert a_decimal("1.234.567,89", D("0.01")) == D("1234567.89")
    assert a_decimal("no soy un número") is None


def test_una_fecha_ilegible_no_se_adivina() -> None:
    assert a_fecha(datetime(2026, 8, 1, 13, 42)) == date(2026, 8, 1)
    assert a_fecha("2026-08-01") == date(2026, 8, 1)
    assert a_fecha("01/08/2026") == date(2026, 8, 1)
    assert a_fecha("el martes") is None


# ── §3.4 Normalización de punta a punta, sobre un libro real ──────────────────


def test_el_co_numerico_y_el_textual_caen_en_el_mismo_punto_de_venta(
    estructura: None, sesion: Session
) -> None:
    """Si no, ALAMEDA 2 saldría dos veces en el reporte y una de ellas huérfana."""
    salida = ingerir(sesion, [fila_venta("606"), fila_venta(606)])

    assert salida.aceptadas == 2
    assert salida.rechazadas == 0
    assert {linea.punto_venta_id for linea in lineas(sesion)} == {id_punto_venta(sesion, "606")}


def test_el_nit_se_guarda_recortado_y_cruza_con_el_catalogo(
    estructura: None, sesion: Session
) -> None:
    """Es lo que habilita el reporte por vendedor, que hoy no existe (§3.4)."""
    salida = ingerir(
        sesion,
        [fila_venta(nit="900983334      ")],
        clientes=[[900983334, "LUXURY HOTELS SAS", "HORECA", "Michel Barrios"]],
    )

    assert salida.aceptadas == 1
    linea = lineas(sesion)[0]
    assert linea.nit_cliente == "900983334"
    cliente = sesion.get(Cliente, linea.cliente_id)
    assert cliente is not None
    assert cliente.vendedor == "Michel Barrios"
    assert cliente.canal == "HORECA"


def test_la_clase_de_cliente_corrupta_entra_sin_clasificar_y_queda_registrada(
    estructura: None, sesion: Session
) -> None:
    """95 907 filas vacías, un nombre de usuario y marcas de tiempo (§3.4)."""
    salida = ingerir(
        sesion,
        [
            fila_venta(clase="002 - CLIENTES NACIONALES"),
            fila_venta(clase=" "),
            fila_venta(clase="johana.muñoz"),
            fila_venta(clase=datetime(2026, 8, 3, 16, 29, 2)),
        ],
    )

    guardadas = [linea.clase_cliente for linea in lineas(sesion)]
    assert guardadas == [
        "002 - CLIENTES NACIONALES",
        "SIN CLASIFICAR",
        "SIN CLASIFICAR",
        "SIN CLASIFICAR",
    ]
    # Las tres degradadas entraron: no son rechazos, la venta es buena.
    assert salida.aceptadas == 4
    assert salida.rechazadas == 0
    # Pero **quedan registradas**: reclasificar en silencio es lo que no vale.
    registrados = motivos(sesion, salida.id, "CLASES DE CLIENTES")
    assert len(registrados) == 3
    assert any(
        "johana.muñoz" in r.valor for r in IngestaService(sesion).rechazos(salida.id) if r.valor
    )


def test_el_domicilio_en_blanco_se_guarda_nulo_y_deja_constancia(
    estructura: None, sesion: Session
) -> None:
    salida = ingerir(
        sesion,
        [fila_venta(domicilio="Si"), fila_venta(domicilio="No"), fila_venta(domicilio=" ")],
    )

    assert [linea.domicilio for linea in lineas(sesion)] == [True, False, None]
    assert motivos(sesion, salida.id, "Domicilio")


# ── §3.1 Categoría por la tabla de mapeo, nunca por un `dict` del código ──────


def test_la_categoria_se_resuelve_por_la_tabla_de_mapeo(estructura: None, sesion: Session) -> None:
    """Incluidas las dos variantes ortográficas de `0006`, que ambas van a OTROS."""
    salida = ingerir(
        sesion,
        [
            fila_venta(categoria="0001 - RES"),
            fila_venta(categoria="0010 - RESTAURANTE"),
            fila_venta(categoria="0006 - QUESO Y LACTEOS"),
            fila_venta(categoria="0006 - QUESOS Y LACTEOS"),
        ],
    )

    nombres = [
        sesion.get(Categoria, linea.categoria_id).nombre  # type: ignore[union-attr]
        for linea in lineas(sesion)
    ]
    assert nombres == ["RES", "ASADERO", "OTROS", "OTROS"]
    # Un mapeo que existe no genera ruido en la bitácora.
    assert not motivos(sesion, salida.id, "CATEGORIA")


def test_la_categoria_sin_mapeo_va_a_otros_y_deja_constancia(
    estructura: None, sesion: Session
) -> None:
    """Es la diferencia entre reclasificar por decisión y por accidente (§3.1)."""
    salida = ingerir(sesion, [fila_venta(categoria="0099 - CAVIAR"), fila_venta(categoria=None)])

    otros = sesion.scalars(select(Categoria).where(Categoria.nombre == "OTROS")).one()
    assert [linea.categoria_id for linea in lineas(sesion)] == [otros.id, otros.id]
    # El texto crudo se conserva junto a la clasificación: sin él no se puede
    # auditar un mapeo mal hecho.
    assert lineas(sesion)[0].categoria_siesa == "0099 - CAVIAR"

    registrados = IngestaService(sesion).rechazos(salida.id)
    sin_mapeo = [r for r in registrados if r.campo == "CATEGORIA"]
    assert len(sin_mapeo) == 2
    assert any(r.valor == "0099 - CAVIAR" for r in sin_mapeo)
    assert salida.rechazadas == 0  # la venta entró; solo se reclasificó


# ── §5 y §7 Idempotencia: la regla crítica ───────────────────────────────────


def test_recargar_el_mismo_rango_reemplaza_y_no_duplica(estructura: None, sesion: Session) -> None:
    """«Reprocesar una fecha reemplaza el día completo; no duplica» (§7)."""
    filas = [
        fila_venta("402", datetime(2026, 8, 3), valor=1000),
        fila_venta("402", datetime(2026, 8, 3), valor=500),
        fila_venta("403", datetime(2026, 8, 4), valor=700),
    ]
    ingerir(sesion, filas)
    ingerir(sesion, filas)
    ingerir(sesion, filas)

    assert sesion.scalar(select(func.count()).select_from(VentaLinea)) == 3
    assert sesion.scalar(select(func.sum(VentaLinea.valor_subtotal))) == D("2200.00")


def test_recargar_un_dia_corregido_lo_reemplaza_entero(estructura: None, sesion: Session) -> None:
    """La segunda carga manda: cuatro líneas malas no sobreviven a una buena."""
    ingerir(sesion, [fila_venta("402", datetime(2026, 8, 3), valor=v) for v in (1, 2, 3, 4)])
    ingerir(sesion, [fila_venta("402", datetime(2026, 8, 3), valor=999)])

    assert [linea.valor_subtotal for linea in lineas(sesion)] == [D("999.00")]


def test_recargar_un_punto_de_venta_no_toca_los_demas(estructura: None, sesion: Session) -> None:
    """El borrado es por `(fecha, punto de venta)`, no por rango.

    Un archivo parcial —tres puntos de venta de dieciséis— no puede llevarse por
    delante los días buenos de los otros trece.
    """
    ingerir(
        sesion,
        [
            fila_venta("402", datetime(2026, 8, 3), valor=100),
            fila_venta("403", datetime(2026, 8, 3), valor=200),
        ],
    )
    ingerir(sesion, [fila_venta("402", datetime(2026, 8, 3), valor=150)])

    valores = {(linea.punto_venta_id, linea.valor_subtotal) for linea in lineas(sesion)}
    assert valores == {
        (id_punto_venta(sesion, "402"), D("150.00")),
        (id_punto_venta(sesion, "403"), D("200.00")),
    }


def test_si_la_carga_falla_a_mitad_no_queda_medio_dia_cargado(
    estructura: None, sesion: Session, monkeypatch: pytest.MonkeyPatch
) -> None:
    """El borrado y la inserción van en la misma transacción (§5).

    Se simula una caída en la fila 3 de la segunda carga. Lo que ya estaba
    cargado tiene que seguir intacto: la alternativa —el día borrado y no
    reescrito— es un reporte que da cero y no avisa.
    """
    ingerir(sesion, [fila_venta("402", datetime(2026, 8, 3), valor=100)])

    original = IngestaService._normalizar
    llamadas = {"n": 0}

    def _explotar(self: IngestaService, *args: Any, **kwargs: Any) -> Any:
        llamadas["n"] += 1
        if llamadas["n"] == 3:
            raise RuntimeError("se cayó la red a mitad de la carga")
        return original(self, *args, **kwargs)

    monkeypatch.setattr(IngestaService, "_normalizar", _explotar)

    salida = ingerir(sesion, [fila_venta("402", datetime(2026, 8, 3), valor=v) for v in (7, 8, 9)])

    assert salida.estado is EstadoCorrida.FALLIDA
    assert salida.aceptadas == 0
    # La carga anterior sobrevive entera: ni duplicada ni a medias.
    assert [linea.valor_subtotal for linea in lineas(sesion)] == [D("100.00")]
    # Y el intento fallido queda en la bitácora con su motivo.
    corrida = sesion.get(CorridaIngesta, salida.id)
    assert corrida is not None
    assert corrida.mensaje is not None
    assert "se cayó la red" in corrida.mensaje


# ── §5 Bitácora: nada se descarta en silencio ────────────────────────────────


def test_una_fila_mala_se_rechaza_con_su_motivo_sin_tumbar_la_corrida(
    estructura: None, sesion: Session
) -> None:
    """Un PDV desconocido y una fecha ilegible no pueden costar el archivo."""
    salida = ingerir(
        sesion,
        [
            fila_venta("402", datetime(2026, 8, 3), valor=1000),
            fila_venta("999", datetime(2026, 8, 3)),  # no existe ese C.O.
            fila_venta("402", "el martes pasado"),  # fecha ilegible
            fila_venta("402", datetime(2026, 8, 3), valor=2000),
        ],
    )

    assert salida.estado is EstadoCorrida.COMPLETADA_CON_RECHAZOS
    assert salida.filas_leidas == 4
    assert salida.aceptadas == 2
    assert salida.rechazadas == 2
    assert salida.filas_leidas == salida.aceptadas + salida.rechazadas

    rechazos = IngestaService(sesion).rechazos(salida.id)
    por_campo = {r.campo: r for r in rechazos}
    assert "999" in (por_campo["C.O."].valor or "")
    assert "punto de venta" in por_campo["C.O."].motivo
    assert por_campo["C.O."].fila == 3  # la fila del Excel, no un índice interno
    assert por_campo["Fecha"].valor == "el martes pasado"
    assert por_campo["Fecha"].fila == 4


def test_la_bitacora_agrega_lo_repetido_en_vez_de_repetirlo(
    estructura: None, sesion: Session
) -> None:
    """En el archivo real son 95 907 filas con el mismo problema (§5).

    Una entrada por fila serían 96 000 registros por corrida y más de un millón
    al año para decir cuatro veces lo mismo. La bitácora tiene que servir para
    leerla, no solo para llenarla.
    """
    salida = ingerir(sesion, [fila_venta(clase=" ") for _ in range(40)])

    assert salida.aceptadas == 40
    entradas = [r for r in IngestaService(sesion).rechazos(salida.id) if r.campo.startswith("CLA")]
    assert len(entradas) == 1
    assert "40 filas" in entradas[0].motivo


def test_un_archivo_entero_malo_no_llena_la_bitacora_de_ruido(
    estructura: None, sesion: Session
) -> None:
    """Los rechazos se detallan hasta un tope y de ahí en adelante se agregan.

    600 filas con un punto de venta inexistente son un problema del archivo, no
    600 problemas distintos.
    """
    salida = ingerir(sesion, [fila_venta("999") for _ in range(600)])

    assert salida.rechazadas == 600
    assert salida.aceptadas == 0
    entradas = IngestaService(sesion).rechazos(salida.id)
    assert len(entradas) < 600
    assert sum(1 for e in entradas if "filas)" in e.motivo) == 1


def test_la_venta_de_un_punto_sin_presupuesto_se_ingiere_igual(
    estructura: None, sesion: Session
) -> None:
    """432 EVENTOS BUCARAMANGA vende y no está presupuestado (§3.1, §7)."""
    salida = ingerir(sesion, [fila_venta("432", datetime(2026, 8, 3), valor=77000)])

    assert salida.aceptadas == 1
    assert salida.rechazadas == 0
    assert lineas(sesion)[0].punto_venta_id == id_punto_venta(sesion, "432")


def test_los_importes_y_los_kilos_se_persisten_como_decimal(
    estructura: None, sesion: Session
) -> None:
    """`Decimal` de extremo a extremo, con la escala de cada medida (§4.5)."""
    ingerir(sesion, [fila_venta(valor=16366867, kilos=370.83, costo=13526912.99, margen=17.35)])

    linea = lineas(sesion)[0]
    assert isinstance(linea.valor_subtotal, Decimal)
    assert linea.valor_subtotal == D("16366867.00")
    assert linea.cantidad_inv == D("370.830")
    assert linea.costo_promedio == D("13526912.99")
    # `MARGEN` se conserva **solo para conciliación** (§4.4).
    assert linea.margen_siesa == D("17.350000")


def test_la_corrida_deja_el_rango_que_traia_el_archivo(estructura: None, sesion: Session) -> None:
    salida = ingerir(
        sesion,
        [
            fila_venta("402", datetime(2026, 8, 5)),
            fila_venta("402", datetime(2026, 8, 1)),
            fila_venta("402", datetime(2026, 8, 9)),
        ],
    )

    assert salida.desde == date(2026, 8, 1)
    assert salida.hasta == date(2026, 8, 9)
    assert salida.duracion_ms is not None


# ── §3.4 Catálogo de clientes ─────────────────────────────────────────────────


def test_el_catalogo_de_clientes_se_lee_sin_fila_de_encabezado(
    estructura: None, sesion: Session
) -> None:
    """La hoja `CLIENTES ` empieza directamente por un cliente."""
    ingerir(
        sesion,
        [fila_venta()],
        clientes=[
            [901733567, "ALAMO EDUCA SAS", "HORECA", "Michel Barrios"],
            [816007746, "FAM SAS", "HORECA", "Eliacid De La Hoz"],
        ],
    )

    clientes = {c.nit: c for c in sesion.scalars(select(Cliente))}
    assert set(clientes) == {"901733567", "816007746"}
    assert clientes["901733567"].razon_social == "ALAMO EDUCA SAS"
    assert clientes["816007746"].vendedor == "Eliacid De La Hoz"


def test_recargar_el_catalogo_actualiza_y_no_duplica(estructura: None, sesion: Session) -> None:
    ingerir(sesion, [fila_venta()], clientes=[[901733567, "ALAMO EDUCA", "HORECA", "Michel"]])
    ingerir(sesion, [fila_venta()], clientes=[[901733567, "ALAMO EDUCA SAS", "HORECA", "Ana"]])

    clientes = list(sesion.scalars(select(Cliente)))
    assert len(clientes) == 1
    assert clientes[0].vendedor == "Ana"
    assert clientes[0].razon_social == "ALAMO EDUCA SAS"


def test_un_libro_sin_hojas_reconocibles_se_rechaza_entero(
    estructura: None, sesion: Session
) -> None:
    """Un 422 con explicación vale más que una corrida vacía en verde."""
    from app.core.errors import ErrorValidacion

    contenido = libro_venta([], con_hoja_venta=False)
    with pytest.raises(ErrorValidacion, match="VENTA"):
        IngestaService(sesion).ingerir_archivo(contenido, "cualquier-cosa.xlsx")


# ── El contrato HTTP (`docs/API.md`, sección Ingesta) ─────────────────────────


def test_la_ingesta_por_http_publica_la_corrida_y_sus_rechazos(
    cliente_http: TestClient, analista: dict[str, str]
) -> None:
    contenido = libro_venta(
        [fila_venta("402", datetime(2026, 8, 3)), fila_venta("999", datetime(2026, 8, 3))]
    )
    respuesta = cliente_http.post(
        "/api/v1/ingesta/archivo",
        files={"archivo": ("venta.xlsx", contenido, _TIPO_XLSX)},
        headers=analista,
    )
    assert respuesta.status_code == 200, respuesta.text
    corrida = respuesta.json()
    assert corrida["estado"] == "COMPLETADA_CON_RECHAZOS"
    assert corrida["aceptadas"] == 1
    assert corrida["rechazadas"] == 1
    assert corrida["quien"] == "analista"

    corridas = cliente_http.get("/api/v1/ingesta/corridas", headers=analista).json()
    assert corridas[0]["id"] == corrida["id"]

    rechazos = cliente_http.get(
        f"/api/v1/ingesta/corridas/{corrida['id']}/rechazos", headers=analista
    ).json()
    assert rechazos[0]["campo"] == "C.O."
    assert rechazos[0]["motivo"]


def test_la_ingesta_alimenta_el_reporte_de_cumplimiento(
    cliente_http: TestClient, analista: dict[str, str], sesion: Session
) -> None:
    """La prueba que cierra el círculo: del `.xlsx` al tablero de gerencia."""
    from tests.conftest import dar_presupuesto

    dar_presupuesto(sesion, "402", "RES", "1000000")
    contenido = libro_venta(
        [
            fila_venta("402", datetime(2026, 8, 3), valor=200000, costo=120000),
            fila_venta("402", datetime(2026, 8, 4), valor=100000, costo=60000),
        ]
    )
    cliente_http.post(
        "/api/v1/ingesta/archivo",
        files={"archivo": ("venta.xlsx", contenido, _TIPO_XLSX)},
        headers=analista,
    )

    cuerpo = cliente_http.get(
        f"/api/v1/reportes/cumplimiento?periodo={PERIODO}", headers=analista
    ).json()
    fila = next(f for f in cuerpo["filas"] if f["punto_venta"] == "402")
    assert fila["venta"] == "300000.00"
    assert fila["cumplimiento"] == "0.3000"


def test_el_rol_de_consulta_no_puede_ingerir(
    cliente_http: TestClient, consulta: dict[str, str]
) -> None:
    respuesta = cliente_http.post(
        "/api/v1/ingesta/archivo",
        files={"archivo": ("venta.xlsx", libro_venta([fila_venta()]), _TIPO_XLSX)},
        headers=consulta,
    )
    assert respuesta.status_code == 403


def test_la_fuente_siesa_sin_token_configurado_dice_que_configurar(
    cliente_http: TestClient, analista: dict[str, str]
) -> None:
    """`FuenteVentaSiesa` ya está implementada, pero necesita credenciales.

    Hasta el 13-ago-2026 esta prueba exigía un 501 «la API de SIESA no se ha
    entregado». Ya se entregó y la fuente la consume —ver
    `tests/test_fuente_siesa.py`—, así que lo que queda por comprobar aquí es lo
    mismo que para la fuente Excel sin ruta: que un fallo de configuración diga
    **qué** configurar en lugar de devolver cero filas en verde. El entorno de
    pruebas no lleva `SIGREP_SIESA_TOKEN`, que es justo ese escenario.
    """
    respuesta = cliente_http.post(
        "/api/v1/ingesta/ejecutar",
        json={"desde": "2026-08-01", "hasta": "2026-08-09", "fuente": "siesa"},
        headers=analista,
    )
    assert respuesta.status_code == 422
    assert "SIGREP_SIESA_TOKEN" in respuesta.json()["detalle"]


def test_la_fuente_excel_sin_ruta_configurada_dice_que_configurar(
    cliente_http: TestClient, analista: dict[str, str]
) -> None:
    """`SIGREP_RUTA_ARCHIVO_VENTA` no tiene valor por defecto a propósito."""
    respuesta = cliente_http.post(
        "/api/v1/ingesta/ejecutar",
        json={"desde": "2026-08-01", "hasta": "2026-08-09", "fuente": "excel"},
        headers=analista,
    )
    assert respuesta.status_code == 422
    assert "SIGREP_RUTA_ARCHIVO_VENTA" in respuesta.json()["detalle"]


def test_salud_publica_la_ultima_ingesta(
    cliente_http: TestClient, analista: dict[str, str]
) -> None:
    cliente_http.post(
        "/api/v1/ingesta/archivo",
        files={"archivo": ("venta.xlsx", libro_venta([fila_venta()]), _TIPO_XLSX)},
        headers=analista,
    )
    assert cliente_http.get("/api/v1/salud").json()["ultima_ingesta"] is not None


# ── §3.3 Carga masiva del presupuesto desde la hoja del negocio ──────────────


def libro_cumplimiento() -> bytes:
    """Reproduce la estructura de la hoja `CUMPLIMIENTO PPTO` del libro real.

    Cinco filas de días hábiles por zona antes del encabezado, y tres clases de
    fila mezcladas debajo: total de grupo, total de punto de venta y presupuesto
    por categoría. Solo la tercera se carga; las otras dos son sumas y cargarlas
    duplicaría el presupuesto de la compañía (§3.3).
    """
    libro = Workbook()
    hoja: Any = libro.active
    hoja.title = "CUMPLIMIENTO PPTO"
    hoja.append([None] * 13)
    hoja.append([None, "BUCARAMANGA Y CENTRO"])
    hoja.append([None, "DIAS HABILES", "DIAS TRABAJOS", "IDEAL"])
    hoja.append([None, 27.5, 7.5, 0.2727])
    hoja.append([None] * 13)
    hoja.append(
        [
            "CODIGO",
            "PDV",
            "PPTO",
            "VENTA AGOSTO",
            "%",
            "PROYECCION",
            "%",
            "VENTA DIARIA",
            "VENTA 2025",
            "CRECIMIENTO",
            "COMISION",
            "Margen",
            "PPTO EN KILO",
        ]
    )
    hoja.append(["001", "GRUPO 1", 1000, 0, 0, 0, 0, 0, 0, 0, None, 0, 500])
    hoja.append([402, "MALAMBO", 1000, 0, 0, 0, 0, 0, 0, 0, None, 0, 500])
    hoja.append([402, "RES", 600, 0, 0, 0, 0, 0, 0, 0, None, 0, 300])
    hoja.append([402, "CERDO", 400, 0, 0, 0, 0, 0, 0, 0, None, 0, 200])
    hoja.append([None, None, 1000, 0, 0, 0, 0, 0, 0, 0, None, 0, 500])  # total general

    memoria = io.BytesIO()
    libro.save(memoria)
    return memoria.getvalue()


def test_la_carga_masiva_lee_la_hoja_cumplimiento_ppto_sin_duplicar_totales(
    cliente_http: TestClient, analista: dict[str, str]
) -> None:
    respuesta = cliente_http.post(
        "/api/v1/presupuesto/carga-masiva",
        files={"archivo": ("ppto.xlsx", libro_cumplimiento(), _TIPO_XLSX)},
        data={"periodo": PERIODO, "motivo": "Carga inicial de agosto"},
        headers=analista,
    )
    assert respuesta.status_code == 200, respuesta.text
    resultado = respuesta.json()
    # Dos filas: RES y CERDO. Ni el total del grupo ni el del punto de venta.
    assert resultado["aceptadas"] == 2
    assert resultado["rechazadas"] == 0

    presupuesto = cliente_http.get(
        f"/api/v1/presupuesto?periodo={PERIODO}", headers=analista
    ).json()
    assert {(p["categoria"], p["monto"]) for p in presupuesto} == {
        ("RES", "600.00"),
        ("CERDO", "400.00"),
    }
    # El presupuesto del punto es la suma de sus categorías: se calcula, no se
    # captura por duplicado. Si se hubieran cargado los totales, sería 2 000.
    cumplimiento = cliente_http.get(
        f"/api/v1/reportes/cumplimiento?periodo={PERIODO}", headers=analista
    ).json()
    fila = next(f for f in cumplimiento["filas"] if f["punto_venta"] == "402")
    assert fila["presupuesto"] == "1000.00"


def test_todos_los_puntos_de_venta_del_archivo_existen_en_el_catalogo(
    estructura: None, sesion: Session
) -> None:
    """Los 16 C.O. medidos en la hoja `VENTA` del archivo real.

    Si mañana SIESA abre un punto nuevo, esta prueba no falla —la ingesta lo
    rechaza con su motivo, que es lo correcto— pero el día que alguien borre un
    punto de venta de la semilla, esto lo dice antes que el reporte.
    """
    medidos = {
        "402",
        "403",
        "405",
        "406",
        "407",
        "409",
        "412",
        "413",
        "414",
        "415",
        "432",
        "603",
        "605",
        "606",
        "701",
        "702",
    }
    catalogados = set(sesion.scalars(select(PuntoVenta.codigo_co)))
    assert medidos <= catalogados


_TIPO_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
