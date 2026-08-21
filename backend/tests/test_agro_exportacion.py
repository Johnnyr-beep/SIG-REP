"""Lo que el libro exportado de agropecuaria tiene que decir por sí solo.

Un archivo de Excel se reenvía por correo, se abre tres semanas después y
sobrevive a la pantalla que lo generó. Todo lo que la pantalla explicaba al lado
del número —qué se dejó fuera, si la lista está recortada— tiene que viajar
dentro del archivo, o quien lo reciba suma una columna, no le cuadra contra el
ERP y no tiene forma de averiguar por qué.

Estas pruebas fijan justamente eso, y una cosa más: que los números salgan como
números. Escribirlos como texto convierte el libro en una imagen de un reporte,
que es lo contrario de lo que pide un negocio que hoy trabaja en Excel.
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook

from app.application.services import agro_exportacion_service as expo
from app.domain.enums import Medida, Semaforo
from app.infrastructure.models.agro_vocabulario import EjeCruce, EjeResumen
from app.schemas.agro import (
    ConciliacionAgro,
    FilaCruceAgro,
    FilaResumenAgro,
    FilaVentaDiariaAgro,
    IndicadoresAgro,
    ParametrosCalculoAgro,
    RespuestaCruceAgro,
    RespuestaResumenAgro,
    RespuestaVentaDiariaAgro,
    TotalesVentaDiariaAgro,
)

CORTE = date(2026, 7, 7)
IMPUESTO = "41476911.00"


def _indicadores(venta: str, **extra: Any) -> dict[str, Any]:
    return {
        "venta": Decimal(venta),
        "venta_valor": Decimal(venta),
        "kilos": Decimal("1000.500"),
        "cantidad": Decimal("120"),
        "lineas_facturadas": 8,
        "participacion": Decimal("0.2885"),
        "semaforo": Semaforo.SIN_PRESUPUESTO,
        **extra,
    }


def _parametros() -> ParametrosCalculoAgro:
    return ParametrosCalculoAgro(
        fecha_corte=CORTE,
        umbrales={"amarillo": "0.9"},
        conciliacion=ConciliacionAgro(
            impuesto_valor=Decimal(IMPUESTO),
            impuesto_kilos=Decimal("0"),
            impuesto_lineas=180,
        ),
    )


def _celdas(contenido: bytes) -> tuple[str, list[Any]]:
    """El texto del libro y sus celdas numéricas, para poder afirmar sobre ambos."""
    hoja = load_workbook(io.BytesIO(contenido)).active
    assert hoja is not None
    texto = "\n".join(
        str(c.value) for fila in hoja.iter_rows() for c in fila if isinstance(c.value, str)
    )
    numeros = [
        c.value for fila in hoja.iter_rows() for c in fila if isinstance(c.value, (int, float))
    ]
    return texto, numeros


def _resumen() -> RespuestaResumenAgro:
    return RespuestaResumenAgro(
        periodo="2026-07",
        fecha_corte=CORTE,
        medida=Medida.VALOR,
        por=EjeResumen.VENDEDOR,
        consolidado=IndicadoresAgro(**_indicadores("3147729924.39")),
        filas=[
            FilaResumenAgro(
                clave="22641887",
                nombre="AFRICANO FRIAS ZULMA ESTER",
                **_indicadores("677660968.91", margen_porcentaje=Decimal("0.0913")),
            ),
            # Margen negativo: la venta entre compañías del grupo sale así de
            # verdad. Si el exportador lo recortara a cero, el libro publicaría
            # una utilidad que nadie tuvo.
            FilaResumenAgro(
                clave="CARNES SANTACRUZ SAS",
                nombre="CARNES SANTACRUZ SAS",
                **_indicadores("345380006.51", margen_porcentaje=Decimal("-0.0437")),
            ),
        ],
        parametros_calculo=_parametros(),
    )


def test_los_importes_van_como_numeros_y_no_como_texto() -> None:
    """El negocio trabaja en Excel: un libro de texto no se puede volver a sumar."""
    _, numeros = _celdas(expo.exportar_resumen(_resumen()))

    assert 3147729924.39 in numeros
    assert 677660968.91 in numeros


def test_el_margen_negativo_se_exporta_negativo() -> None:
    """−4,37 % es la venta a una compañía del grupo. Recortarla a cero mentiría."""
    _, numeros = _celdas(expo.exportar_resumen(_resumen()))

    assert -0.0437 in numeros


def test_el_libro_dice_cuanto_impuesto_se_dejo_fuera() -> None:
    """Sin esta nota el archivo no cuadra contra el ERP y nadie sabe por qué."""
    texto, numeros = _celdas(expo.exportar_resumen(_resumen()))

    assert "Impuesto excluido" in texto
    assert 41476911.00 in numeros
    assert "recaudo a nombre de terceros" in texto


def test_la_fecha_de_corte_va_en_la_primera_pagina() -> None:
    """§6: un reporte sin fecha de corte es un reporte que alguien malinterpreta."""
    texto, _ = _celdas(expo.exportar_resumen(_resumen()))

    assert f"Fecha de corte: {CORTE}" in texto


def _cruce(*, truncado: bool) -> RespuestaCruceAgro:
    return RespuestaCruceAgro(
        periodo="2026-07",
        fecha_corte=CORTE,
        medida=Medida.VALOR,
        por=EjeCruce.VENDEDOR_CLIENTE_PRODUCTO,
        ejes=["vendedor", "cliente", "item"],
        consolidado=IndicadoresAgro(**_indicadores("3147729924.39")),
        filas=[
            FilaCruceAgro(
                claves=["830505537", "INVERSIONES Y NEGOCIOS", "CANAL"],
                nombres=["AGROPECUARIA SANTACRUZ LTDA", "INVERSIONES Y NEGOCIOS", "CANAL DE CERDO"],
                **_indicadores("104972610.00"),
            )
        ],
        truncado=truncado,
        limite=500,
        parametros_calculo=_parametros(),
    )


def test_un_cruce_truncado_lo_dice_dentro_del_archivo() -> None:
    """Con el límite en 500, en una carga real quedan 198 millones fuera de la vista.

    El aviso no puede vivir solo en la pantalla: el archivo se reenvía y quien lo
    abra va a sumar la columna y a compararla con el consolidado.
    """
    texto, _ = _celdas(expo.exportar_cruce(_cruce(truncado=True)))

    assert "filas de mayor venta" in texto
    assert "500" in texto


def test_un_cruce_completo_no_lleva_ese_aviso() -> None:
    """Advertir de un recorte que no hubo enseña a ignorar el aviso."""
    texto, _ = _celdas(expo.exportar_cruce(_cruce(truncado=False)))

    assert "filas de mayor venta" not in texto


def test_el_cruce_pinta_una_columna_por_eje() -> None:
    """Los dos cruces tienen distinto número de ejes; la cabecera sale de `ejes`."""
    texto, _ = _celdas(expo.exportar_cruce(_cruce(truncado=False)))

    assert "Vendedor" in texto
    assert "Cliente" in texto
    assert "CANAL DE CERDO" in texto


def test_la_fila_de_totales_es_la_que_publica_la_respuesta() -> None:
    """No se recalcula en el exportador: así el archivo no puede discrepar.

    Los valores se afirman uno a uno, no su suma: una suma correcta con las
    columnas cambiadas de sitio también daría el mismo total.
    """
    datos = RespuestaVentaDiariaAgro(
        periodo="2026-07",
        fecha_corte=CORTE,
        desde=date(2026, 7, 1),
        hasta=date(2026, 7, 2),
        medida=Medida.VALOR,
        fechas=[date(2026, 7, 1), date(2026, 7, 2)],
        presupuesto_diario_por_centro={"301": None, "302": None},
        filas=[
            FilaVentaDiariaAgro(
                centro="301",
                nombre="AGROPECUARIA SANTACRUZ LTDA",
                valores=[Decimal("544946594.35"), Decimal("455688945.71")],
                total=Decimal("1000635540.06"),
            )
        ],
        totales=TotalesVentaDiariaAgro(
            valores=[Decimal("544946594.35"), Decimal("455688945.71")],
            total=Decimal("1000635540.06"),
            presupuesto_diario=None,
        ),
        parametros_calculo=_parametros(),
    )

    hoja = load_workbook(io.BytesIO(expo.exportar_venta_diaria(datos))).active
    assert hoja is not None
    fila = next(f for f in hoja.iter_rows(values_only=True) if str(f[1] or "").startswith("TOTAL"))

    assert fila[3] == 544946594.35
    assert fila[4] == 455688945.71
    assert fila[5] == 1000635540.06


def test_el_presupuesto_diario_sin_capturar_deja_la_celda_vacia() -> None:
    """`None` es «no hay meta», que no es «la meta es cero» (§7)."""
    datos = RespuestaVentaDiariaAgro(
        periodo="2026-07",
        fecha_corte=CORTE,
        desde=date(2026, 7, 1),
        hasta=date(2026, 7, 1),
        medida=Medida.VALOR,
        fechas=[date(2026, 7, 1)],
        presupuesto_diario_por_centro={"301": None},
        filas=[
            FilaVentaDiariaAgro(
                centro="301",
                nombre="AGROPECUARIA SANTACRUZ LTDA",
                valores=[Decimal("544946594.35")],
                total=Decimal("544946594.35"),
            )
        ],
        totales=TotalesVentaDiariaAgro(
            valores=[Decimal("544946594.35")],
            total=Decimal("544946594.35"),
            presupuesto_diario=None,
        ),
        parametros_calculo=_parametros(),
    )

    hoja = load_workbook(io.BytesIO(expo.exportar_venta_diaria(datos))).active
    assert hoja is not None
    fila = next(f for f in hoja.iter_rows(values_only=True) if f[0] == "301")

    assert fila[2] is None
