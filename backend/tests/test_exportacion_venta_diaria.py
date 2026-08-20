"""La exportación de venta diaria dice lo mismo que la pantalla.

Dos invariantes, y las dos existen por el mismo motivo: **el archivo no puede
discrepar de la pantalla**. Quien abra el Excel no tiene forma de saber cuál de
las dos versiones vale, así que no puede haber dos.

1. La fila de totales se **exporta**, no se recalcula aquí.
2. Cuando el rango cruza de mes va **una columna de referencia por período**.
   Escribir una sola publicaria la referencia equivocada para la mitad de las
   columnas, y promediarlas inventaria un número que no existe en ningún sitio.
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from app.application.services.exportacion_service import exportar_venta_diaria
from app.domain.enums import Medida
from app.schemas.reportes import (
    FilaVentaDiaria,
    ParametrosCalculo,
    RespuestaVentaDiaria,
    TotalesVentaDiaria,
)


def _respuesta(*, periodos: list[str], fechas: list[date]) -> RespuestaVentaDiaria:
    return RespuestaVentaDiaria(
        periodo=periodos[-1],
        fecha_corte=fechas[-1],
        desde=fechas[0],
        hasta=fechas[-1],
        medida=Medida.VALOR,
        periodos=periodos,
        fechas=fechas,
        presupuesto_diario_por_pdv={"402": "1000.00"},
        presupuesto_diario_por_periodo={
            periodo: {"402": f"{(indice + 1) * 1000}.00"} for indice, periodo in enumerate(periodos)
        },
        filas=[
            FilaVentaDiaria(
                punto_venta="402",
                nombre="MALAMBO",
                valores=["100.00"] * len(fechas),
                total=f"{100 * len(fechas)}.00",
            )
        ],
        totales=TotalesVentaDiaria(
            valores=["100.00"] * len(fechas),
            total=f"{100 * len(fechas)}.00",
            presupuesto_diario="1000.00",
            presupuesto_diario_por_periodo={
                periodo: f"{(indice + 1) * 1000}.00" for indice, periodo in enumerate(periodos)
            },
        ),
        parametros_calculo=ParametrosCalculo(fecha_corte=fechas[-1], umbrales={}),
    )


def _hoja(datos: RespuestaVentaDiaria):
    return load_workbook(io.BytesIO(exportar_venta_diaria(datos))).active


def test_la_fila_de_totales_se_exporta() -> None:
    hoja = _hoja(_respuesta(periodos=["2026-08"], fechas=[date(2026, 8, 1), date(2026, 8, 2)]))
    ultima = [c.value for c in hoja[hoja.max_row]]

    assert "TOTAL" in str(ultima[1])
    # El total del archivo es el que publica la respuesta, no uno recalculado.
    # Se compara como numero y no como texto a proposito: el export escribe
    # numeros para que quien reciba el archivo pueda seguir operando con el, asi
    # que «200» y «200.00» son el mismo valor y solo difieren en el formato.
    assert Decimal(str(ultima[-1])) == Decimal("200")


def test_un_solo_periodo_conserva_una_sola_columna_de_referencia() -> None:
    """El modo de siempre no cambia: quien exporte un mes ve lo de antes."""
    hoja = _hoja(_respuesta(periodos=["2026-08"], fechas=[date(2026, 8, 1)]))
    encabezado = [c.value for c in hoja[4]]

    assert encabezado[:3] == ["Punto de venta", "Nombre", "Ppto. diario"]


def test_un_rango_que_cruza_de_mes_lleva_una_referencia_por_periodo() -> None:
    """Un día de julio no se mide contra el presupuesto de agosto."""
    hoja = _hoja(
        _respuesta(periodos=["2026-07", "2026-08"], fechas=[date(2026, 7, 31), date(2026, 8, 1)])
    )
    encabezado = [c.value for c in hoja[4]]

    assert encabezado[2] == "Ppto. diario 2026-07"
    assert encabezado[3] == "Ppto. diario 2026-08"

    # Y cada una lleva su propio valor, no el mismo repetido.
    fila = [c.value for c in hoja[5]]
    assert Decimal(str(fila[2])) == Decimal("1000")
    assert Decimal(str(fila[3])) == Decimal("2000")
